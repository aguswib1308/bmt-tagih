from werkzeug.utils import secure_filename
from flask import Flask, request, jsonify, session, render_template, redirect, url_for, send_from_directory, make_response
import sqlite3
import hashlib
import requests
import os
import hmac
import re
import io
from PIL import Image
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload

GDRIVE_FOLDER_ID = "1ftwqyYVBIu2YZnjLip36vylMeB3dPvLV"
GDRIVE_CREDS_PATH = os.path.join("data", "gdrive_credentials.json")

def compress_image(file_obj, max_bytes=1*1024*1024):
    img = Image.open(file_obj)
    if img.mode in ("RGBA", "P", "CMYK"):
        img = img.convert("RGB")
    max_dim = 1920
    if max(img.size) > max_dim:
        ratio = max_dim / max(img.size)
        img = img.resize((int(img.width*ratio), int(img.height*ratio)), Image.LANCZOS)
    for quality in [85, 70, 55, 40, 25]:
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=quality, optimize=True)
        if buf.tell() <= max_bytes:
            buf.seek(0)
            return buf
    buf.seek(0)
    return buf

def get_gdrive_creds():
    import json
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    from datetime import datetime
    SCOPES = ["https://www.googleapis.com/auth/drive"]
    token_path = os.path.join("data", "gdrive_token.json")
    with open(token_path) as f:
        tdata = json.load(f)
    expiry = None
    if tdata.get("expiry"):
        try:
            expiry = datetime.fromisoformat(tdata["expiry"])
        except Exception:
            expiry = None
    creds = Credentials(
        token=tdata.get("token"),
        refresh_token=tdata.get("refresh_token"),
        token_uri=tdata.get("token_uri", "https://oauth2.googleapis.com/token"),
        client_id=tdata.get("client_id"),
        client_secret=tdata.get("client_secret"),
        scopes=tdata.get("scopes", SCOPES),
        expiry=expiry
    )
    if not creds.valid or creds.expiry is None:
        creds.refresh(Request())
        tdata["token"] = creds.token
        if creds.expiry:
            tdata["expiry"] = creds.expiry.isoformat()
        with open(token_path, "w") as f:
            json.dump(tdata, f)
    return creds

def upload_to_gdrive(image_buf, filename):
    try:
        creds = get_gdrive_creds()
        svc = build("drive", "v3", credentials=creds)
        clean_name = filename.rsplit(".", 1)[0] + ".jpg"
        meta = {"name": clean_name, "parents": [GDRIVE_FOLDER_ID]}
        media = MediaIoBaseUpload(image_buf, mimetype="image/jpeg", resumable=False)
        f = svc.files().create(body=meta, media_body=media, fields="id").execute()
        file_id = f.get("id")
        svc.permissions().create(
            fileId=file_id, body={"type": "anyone", "role": "reader"}
        ).execute()
        return file_id
    except Exception as e:
        print("Drive upload error:", e)
        return None

from datetime import datetime, timedelta
from functools import wraps

app = Flask(__name__)

from keanggotaan_routes import kean_bp
app.register_blueprint(kean_bp)
_sk = os.environ.get("SECRET_KEY")
if not _sk:
    _sk_file = os.path.join("data", ".secret_key")
    if os.path.exists(_sk_file):
        with open(_sk_file) as _f: _sk = _f.read().strip()
    else:
        import secrets as _sec
        _sk = _sec.token_hex(32)
        os.makedirs("data", exist_ok=True)
        with open(_sk_file, "w") as _f: _f.write(_sk)
        print("[SECURITY] Secret key baru digenerate:", _sk_file)
app.secret_key = _sk
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10MB

# â”€â”€ Auto-create folder data/ saat pertama jalan â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
os.makedirs("data", exist_ok=True)
os.makedirs("data/foto_kunjungan", exist_ok=True)

DB_PATH = os.environ.get("DB_PATH", "data/koperasi.db")
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)
app.config['SESSION_PERMANENT'] = True
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax' 

# â”€â”€ Fonnte Config â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
FONNTE_TOKEN = os.environ.get("FONNTE_TOKEN", "")
NOTIF_AKTIF = os.environ.get("NOTIF_AKTIF", "1")

# â”€â”€ Helper â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def format_rp(n):
    return f"Rp {int(n):,}".replace(",", ".")

def format_tgl_jt(tgl_str):
    if not tgl_str: return "-"
    tgl_str = str(tgl_str).strip()
    if re.match(r'^\d{8}$', tgl_str):
        return tgl_str[-2:]
    if re.match(r'^\d{4}-\d{2}-\d{2}', tgl_str):
        return tgl_str.split('-')[2][:2]
    if re.match(r'^\d{1,2}[/-]\d{1,2}', tgl_str):
        return re.split(r'[/-]', tgl_str)[0]
    m = re.search(r'^(\d{1,2})\b', tgl_str)
    if m: return m.group(1)
    m = re.search(r'\b(\d{1,2})$', tgl_str)
    if m: return m.group(1)
    return tgl_str

def get_setting(key, default=None):
    """Ambil satu nilai dari tabel app_settings."""
    try:
        conn = get_db()
        conn.execute("""
            CREATE TABLE IF NOT EXISTS app_settings (
                key TEXT PRIMARY KEY,
                value TEXT,
                keterangan TEXT
            )
        """)
        conn.commit()
        row = conn.execute("SELECT value FROM app_settings WHERE key=?", (key,)).fetchone()
        conn.close()
        return row["value"] if row else default
    except:
        return default


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        return f(*args, **kwargs)
    return decorated

def get_user_role():
    return session.get("role", "marketing")

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        if get_user_role() != "admin":
            return jsonify({"error": "Admin only"}), 403
        return f(*args, **kwargs)
    return decorated

# â”€â”€ Kirim WA via Fonnte â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def kirim_wa(no_hp, pesan):
    flag_path = os.path.join("data", "notif_flag.txt")
    try:
        with open(flag_path, "r") as ff:
            flag = ff.read().strip()
    except:
        flag = os.environ.get("NOTIF_AKTIF", "0")
    if flag != "1":
        return {"status": "skip", "reason": "Notifikasi dinonaktifkan"}
    if not FONNTE_TOKEN:
        return {"status": "skip", "reason": "FONNTE_TOKEN belum diset"}
    try:
        no_hp = no_hp.replace("-", "").replace(" ", "")
        if no_hp.startswith("0"):
            no_hp = "62" + no_hp[1:]
        r = requests.post(
            "https://api.fonnte.com/send",
            headers={"Authorization": FONNTE_TOKEN},
            data={"target": no_hp, "message": pesan},
            timeout=10
        )
        return r.json()
    except Exception as e:
        return {"error": str(e)}

def get_template_isi(template_id):
    try:
        conn = get_db()
        row = conn.execute("SELECT isi FROM template_pesan WHERE id=?", (template_id,)).fetchone()
        conn.close()
        return row["isi"] if row else None
    except:
        return None

def pesan_tagihan(nasabah_nama, total, jatuh_tempo, marketing_nama, no_akad="", tunggakan=0):
    tunggakan = tunggakan or 0
    if tunggakan > 0:
        template = get_template_isi("tagihan_tunggakan")
        if template:
            return template.format(
                nasabah_nama=nasabah_nama,
                total=format_rp(total),
                jatuh_tempo=jatuh_tempo,
                no_akad=no_akad,
                tunggakan=format_rp(tunggakan),
                total_keseluruhan=format_rp(total + tunggakan)
            )
    template = get_template_isi("tagihan")
    if template:
        return template.format(
            nasabah_nama=nasabah_nama,
            total=format_rp(total),
            jatuh_tempo=jatuh_tempo,
            no_akad=no_akad,
            marketing_nama=marketing_nama
        )
    return f"Assalamu'alaikum, {nasabah_nama}\n\nTagihan: {format_rp(total)}\nJatuh Tempo: {jatuh_tempo}"

def pesan_lunas(nasabah_nama, jumlah, marketing_nama):
    template = get_template_isi("lunas")
    if template:
        return template.format(
            nasabah_nama=nasabah_nama,
            jumlah=format_rp(jumlah),
            tgl_sekarang=datetime.now().strftime('%d/%m/%Y %H:%M'),
            marketing_nama=marketing_nama
        )
    return f"""Assalamu'alaikum, {nasabah_nama} 🙏\n\nPembayaran {format_rp(jumlah)} berhasil dicatat.\nMarketing: {marketing_nama}"""

# â”€â”€ Health Check (Railway butuh ini) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/health")
def health():
    return jsonify({"status": "ok", "time": datetime.now().isoformat()})

# â”€â”€ AUTH â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/settings", methods=["GET"])
@admin_required
def get_all_settings():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY, value TEXT, keterangan TEXT
        )
    """)
    defaults = [
        ("delay_blast_detik", "10", "Jeda antar pesan WA saat blast (detik)"),
    ]
    for k, v, ket in defaults:
        conn.execute("INSERT OR IGNORE INTO app_settings (key,value,keterangan) VALUES (?,?,?)",
                     (k, v, ket))
    conn.commit()
    rows = conn.execute("SELECT key, value, keterangan FROM app_settings ORDER BY key").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/settings/<key>", methods=["PUT"])
@admin_required
def update_setting(key):
    data = request.json
    value = str(data.get("value", ""))
    conn = get_db()
    conn.execute("""
        INSERT INTO app_settings (key, value) VALUES (?,?)
        ON CONFLICT(key) DO UPDATE SET value=excluded.value
    """, (key, value))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/login", methods=["POST"])
def login():
    data = request.json
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()

    conn = get_db()
    user = conn.execute(
        "SELECT * FROM users WHERE username=? AND password=? AND aktif=1",
        (username, hash_pw(password))
    ).fetchone()
    conn.close()

    if not user:
        return jsonify({"error": "Username atau password salah"}), 401

    session.permanent = True
    session["user_id"]     = user["id"]
    session["username"]    = user["username"]
    session["nama"]        = user["nama"]
    session["role"]        = user["role"]
    session["marketing_id"]= user["marketing_id"]

    return jsonify({
        "success": True,
        "nama": user["nama"],
        "role": user["role"],
        "marketing_id": user["marketing_id"]
    })

@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"success": True})

@app.route("/api/me")
def me():
    if "user_id" not in session:
        return jsonify({"login": False})
    return jsonify({
        "login": True,
        "nama": session.get("nama"),
        "username": session.get("username"),
        "role": get_user_role(),
        "marketing_id": session.get("marketing_id")
    })

@app.route("/api/me/password", methods=["POST"])
@login_required
def change_my_password():
    """Ganti password sendiri (self-service, semua role)."""
    data = request.json or {}
    old_pw = (data.get("old_password") or "").strip()
    new_pw = (data.get("new_password") or "").strip()
    if len(new_pw) < 4:
        return jsonify({"error": "Password baru minimal 4 karakter"}), 400
    uid = session.get("user_id")
    if not uid:
        return jsonify({"error": "Sesi tidak valid"}), 401
    conn = get_db()
    row = conn.execute("SELECT id, password FROM users WHERE id=?", (uid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "User tidak ditemukan"}), 404
    if row["password"] != hash_pw(old_pw):
        conn.close()
        return jsonify({"error": "Password lama salah"}), 400
    conn.execute("UPDATE users SET password=? WHERE id=?", (hash_pw(new_pw), uid))
    conn.commit()
    conn.close()
    return jsonify({"success": True})

# â”€â”€ USER MANAGEMENT (ADMIN) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/users", methods=["GET"])
@admin_required
def get_users():
    conn = get_db()
    rows = conn.execute("SELECT id, username, nama, role, marketing_id, aktif FROM users ORDER BY role, nama").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/users", methods=["POST"])
@admin_required
def create_user():
    data = request.json
    username = data.get("username", "").strip()
    nama = data.get("nama", "").strip()
    marketing_id = data.get("marketing_id", "").strip()
    role = data.get("role", "marketing")

    if not username or not nama:
        return jsonify({"error": "Username dan nama wajib diisi"}), 400

    conn = get_db()
    exist = conn.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone()
    if exist:
        conn.close()
        return jsonify({"error": "Username sudah dipakai"}), 400

    conn.execute(
        "INSERT INTO users (username, password, nama, role, marketing_id, aktif) VALUES (?, ?, ?, ?, ?, 1)",
        (username, hash_pw("bmt2026"), nama, role, marketing_id)
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route("/api/users/<int:user_id>", methods=["PUT"])
@admin_required
def update_user(user_id):
    data = request.json
    username = data.get("username", "").strip()
    nama = data.get("nama", "").strip()
    marketing_id = data.get("marketing_id", "").strip()
    aktif = int(data.get("aktif", 1))

    if not username or not nama:
        return jsonify({"error": "Username dan nama wajib diisi"}), 400

    conn = get_db()
    exist = conn.execute("SELECT id FROM users WHERE username=? AND id!=?", (username, user_id)).fetchone()
    if exist:
        conn.close()
        return jsonify({"error": "Username sudah dipakai user lain"}), 400

    conn.execute(
        "UPDATE users SET username=?, nama=?, marketing_id=?, aktif=? WHERE id=?",
        (username, nama, marketing_id, aktif, user_id)
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route("/api/users/<int:user_id>/password", methods=["PUT"])
@admin_required
def change_user_password(user_id):
    data = request.json
    new_pw = (data.get("password") or "").strip()
    if len(new_pw) < 4:
        return jsonify({"error": "Password minimal 4 karakter"}), 400
    conn = get_db()
    user = conn.execute("SELECT id FROM users WHERE id=?", (user_id,)).fetchone()
    if not user:
        conn.close()
        return jsonify({"error": "User tidak ditemukan"}), 404
    conn.execute("UPDATE users SET password=? WHERE id=?", (hash_pw(new_pw), user_id))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/users/<int:user_id>/reset", methods=["PUT"])
@admin_required
def reset_user_password(user_id):
    conn = get_db()
    conn.execute("UPDATE users SET password=? WHERE id=?", (hash_pw("bmt2026"), user_id))
    conn.commit()
    conn.close()
    return jsonify({"success": True})

# â”€â”€ DASHBOARD â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/dashboard")
@login_required
def dashboard():
    conn = get_db()
    bulan = request.args.get("bulan", datetime.now().strftime("%Y-%m"))
    marketing_id = session.get("marketing_id")
    role = get_user_role()

    where = "WHERE t.bulan=?"
    params = [bulan]
    if role not in ("admin","leader","petugas") and marketing_id:
        where += " AND n.marketing_id=?"
        params.append(marketing_id)

    stats = conn.execute(f"""
        SELECT
            COUNT(*) as total_nasabah,
            SUM(t.angsuran_per_bulan) as total_tagihan,
            SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN t.angsuran_per_bulan ELSE 0 END) as total_terkumpul,
            SUM(CASE WHEN t.status='BELUM' AND t.total_tagihan >= 1 THEN t.total_tagihan ELSE 0 END) as total_tunggakan,
            SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN 1 ELSE 0 END) as sudah_bayar,
            SUM(CASE WHEN (t.status='BELUM') THEN 1 ELSE 0 END) as belum_bayar
        FROM tagihan t
        JOIN nasabah n ON t.no_rekening = n.no_rekening
        {where}
    """, params).fetchone()

    rekap_marketing = []
    if role in ("admin","leader"):
        rows = conn.execute(f"""
            SELECT n.marketing_nama,
                COUNT(*) as total,
                SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN 1 ELSE 0 END) as lunas,
                SUM(CASE WHEN (t.status='BELUM') THEN 1 ELSE 0 END) as belum,
                SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN t.angsuran_per_bulan ELSE 0 END) as nominal_lunas
            FROM tagihan t
            JOIN nasabah n ON t.no_rekening = n.no_rekening
            WHERE t.bulan=?
              AND n.marketing_nama IS NOT NULL
              AND n.marketing_nama NOT GLOB '[0-9]*'
            GROUP BY n.marketing_nama
            ORDER BY lunas DESC
        """, [bulan]).fetchall()
        rekap_marketing = [dict(r) for r in rows]

    conn.close()
    return jsonify({
        "stats": dict(stats),
        "rekap_marketing": rekap_marketing,
        "bulan": bulan
    })

# â”€â”€ TAGIHAN LIST â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/tagihan")
@login_required
def list_tagihan():
    conn = get_db()
    bulan        = request.args.get("bulan", datetime.now().strftime("%Y-%m"))
    status       = request.args.get("status", "")
    search       = request.args.get("q", "")
    marketing_id = session.get("marketing_id")
    role         = get_user_role()

    where  = ["t.bulan=?"]
    params = [bulan]

    if role not in ("admin","leader","petugas") and marketing_id:
        where.append("n.marketing_id=?")
        params.append(marketing_id)
    if status == "LUNAS":
        where.append("(t.status IN ('LUNAS','SUDAH_BAYAR'))")
    elif status == "BELUM":
        where.append("t.status='BELUM'")
        where.append("t.total_tagihan >= 1")
    elif status:
        where.append("t.status=?")
        params.append(status)
    kolek = request.args.get("kolek", "")
    if kolek:
        where.append("t.kolektibilitas=?")
        params.append(int(kolek))
    reschedule = request.args.get("reschedule", "")
    if reschedule == "1":
        where.append("n.is_reschedule=1")
    elif reschedule == "0":
        where.append("(n.is_reschedule=0 OR n.is_reschedule IS NULL)")
    if search:
        where.append("(n.nama LIKE ? OR n.no_rekening LIKE ?)")
        params += [f"%{search}%", f"%{search}%"]

    sql = f"""
        SELECT t.id, t.no_rekening, n.nama, n.no_hp, n.marketing_nama,
               t.saldo_pinjaman, t.tunggakan_pokok, t.tunggakan_margin,
               t.total_tagihan, t.kolektibilitas, t.status, t.keterangan,
               t.cara_bayar, t.tgl_angsuran, n.tanggal_jt, n.alamat, n.is_reschedule
        FROM tagihan t
        JOIN nasabah n ON t.no_rekening = n.no_rekening
        WHERE {' AND '.join(where)}
        ORDER BY t.kolektibilitas DESC, n.tanggal_jt ASC, t.total_tagihan DESC
    """
    # Pagination
    limit  = min(int(request.args.get("limit", 50)), 500)
    offset = int(request.args.get("offset", 0))

    sql_count = f"SELECT COUNT(*) FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening WHERE {' AND '.join(where)}"
    total = conn.execute(sql_count, params).fetchone()[0]

    params.append(limit)
    params.append(offset)
    sql += " LIMIT ? OFFSET ?"
    rows = conn.execute(sql, params).fetchall()

    # Jumlah kunjungan per nasabah bulan ini (untuk badge di marketing & petugas)
    ensure_kunjungan_table(conn)
    kj = conn.execute(
        "SELECT no_rekening, COUNT(*) as jumlah FROM kunjungan WHERE bulan=? GROUP BY no_rekening",
        [bulan]).fetchall()
    kj_map = {r["no_rekening"]: r["jumlah"] for r in kj}
    conn.close()
    data = []
    for r in rows:
        d = dict(r)
        d["jumlah_kunjungan"] = kj_map.get(r["no_rekening"], 0)
        data.append(d)
    return jsonify({
        "data": data,
        "total": total,
        "limit": limit,
        "offset": offset
    })

# â”€â”€ CATAT PEMBAYARAN â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.route("/api/tagihan/jatuh-tempo")
@login_required
def tagihan_jatuh_tempo():
    from datetime import datetime
    bulan = request.args.get("bulan", datetime.now().strftime("%Y-%m"))
    wa_mode = request.args.get("wa") == "1"
    today_day = int(datetime.now().strftime("%d"))
    conn = get_db()
    marketing_id = session.get("marketing_id")
    role = get_user_role()
    extra = ""
    extra_params = []
    if role not in ("admin", "leader", "petugas") and marketing_id:
        extra = " AND n.marketing_id=?"
        extra_params.append(marketing_id)

    if wa_mode:
        # Mode WA penagihan: window 4 hari (hari ini + 3 hari ke belakang),
        # hanya kolektibilitas lancar(1) & DPK(2), kecuali reschedule.
        # wa_count = berapa kali sudah dikirim WA bulan ini (untuk badge).
        ensure_wa_send_log(conn)
        start_day = max(1, today_day - 3)
        params = [bulan, bulan, start_day, today_day] + extra_params
        rows = conn.execute(f"""
            SELECT t.id, t.no_rekening, t.total_tagihan, t.status, t.kolektibilitas,
                   t.tunggakan_pokok, t.tunggakan_margin, t.cara_bayar, t.angsuran_per_bulan,
                   n.nama, n.no_hp, n.tanggal_jt, n.marketing_nama, n.alamat,
                   n.is_reschedule,
                   (SELECT COUNT(*) FROM wa_send_log w
                     WHERE w.no_rekening=t.no_rekening AND w.bulan=? AND w.status='terkirim') as wa_count
            FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
            WHERE t.bulan=? AND t.status='BELUM'
              AND t.total_tagihan >= 1
              AND t.kolektibilitas IN (1,2)
              AND (n.is_reschedule=0 OR n.is_reschedule IS NULL)
              AND CAST(SUBSTR(n.tanggal_jt, 9, 2) AS INTEGER) BETWEEN ? AND ?
              {extra}
            ORDER BY (CASE WHEN (SELECT COUNT(*) FROM wa_send_log w2
                     WHERE w2.no_rekening=t.no_rekening AND w2.bulan=? AND w2.status='terkirim')>0 THEN 1 ELSE 0 END),
                     CAST(SUBSTR(n.tanggal_jt, 9, 2) AS INTEGER), n.nama
        """, params + [bulan]).fetchall()
        data = []
        for r in rows:
            d = dict(r)
            d["sudah_wa"] = (d.get("wa_count") or 0) > 0
            data.append(d)
        conn.close()
        return jsonify({"data": data, "total": len(data),
                        "today_day": today_day, "start_day": start_day,
                        "window": 4, "bulan": bulan})

    # Mode default (kumulatif s/d hari ini) - dipakai filter halaman Tagihan
    params = [bulan, today_day] + extra_params
    rows = conn.execute(f"""
        SELECT t.id, t.no_rekening, t.total_tagihan, t.status, t.kolektibilitas,
               t.tunggakan_pokok, t.tunggakan_margin, t.cara_bayar,
               n.nama, n.no_hp, n.tanggal_jt, n.marketing_nama, n.alamat,
               n.is_reschedule
        FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
        WHERE t.bulan=? AND t.status='BELUM'
          AND t.total_tagihan >= 1
          AND CAST(SUBSTR(n.tanggal_jt, 9, 2) AS INTEGER) <= ?
          {extra}
        ORDER BY CAST(SUBSTR(n.tanggal_jt, 9, 2) AS INTEGER), n.nama
    """, params).fetchall()
    conn.close()
    return jsonify({"data": [dict(r) for r in rows], "total": len(rows),
                    "today_day": today_day, "bulan": bulan})


def _run_sync_mespro(timeout=120):
    """Jalankan sync_mespro.py sebagai subprocess. Return dict hasil.
    Dipakai endpoint sync manual & auto-sync sebelum blast WA.
    Bila gagal (mis. tunnel MESPro mati), kembalikan ok=False -- caller
    boleh memutuskan tetap lanjut pakai data terakhir."""
    import subprocess, re as _re
    try:
        result = subprocess.run(
            [os.path.join(os.path.dirname(__file__), "venv", "bin", "python3"),
             os.path.join(os.path.dirname(__file__), "sync_mespro.py")],
            capture_output=True, text=True, timeout=timeout,
            cwd=os.path.dirname(__file__)
        )
        output = result.stdout + result.stderr
        if result.returncode != 0:
            return {"ok": False, "error": "Sync gagal", "detail": output[-500:]}
        m_baru  = _re.search(r"Baru\s*:\s*(\d+)", output)
        m_upd   = _re.search(r"Update\s*:\s*(\d+)", output)
        m_skip  = _re.search(r"Skip\s*:\s*(\d+)", output)
        m_query = _re.search(r"(\d+) nasabah kredit aktif", output)
        return {
            "ok": True,
            "total_query": int(m_query.group(1)) if m_query else 0,
            "baru": int(m_baru.group(1)) if m_baru else 0,
            "update": int(m_upd.group(1)) if m_upd else 0,
            "skip": int(m_skip.group(1)) if m_skip else 0,
        }
    except subprocess.TimeoutExpired:
        return {"ok": False, "error": "Timeout - sync terlalu lama (>%d detik)" % timeout}
    except Exception as e:
        return {"ok": False, "error": str(e)}

def _sync_msg(s):
    """Pesan ringkas status sync utk ditampilkan ke admin di hasil blast."""
    if s.get("ok"):
        return "Sync MESPro OK: %d update, %d baru (data terbaru)" % (
            s.get("update", 0), s.get("baru", 0))
    return "Sync dilewati, blast pakai data terakhir (%s)" % str(s.get("error", ""))[:120]

@app.route("/api/reminder/blast-jt-hari-ini", methods=["POST"])
@admin_required
def blast_jt_hari_ini():
    """Kirim notif WA ke semua anggota yang jatuh tempo hari ini.
    Sync MESPro dijalankan dulu agar total tagihan memuat angsuran bulan
    berjalan (cegah WA kurang 1 bulan saat MESPro baru menggulung tunggakan)."""
    from datetime import datetime
    bulan = request.json.get("bulan", datetime.now().strftime("%Y-%m"))
    today_day = int(datetime.now().strftime("%d"))
    start_day = max(1, today_day - 3)
    # Auto-sync sebelum blast (tetap lanjut walau gagal)
    sync_info = _run_sync_mespro()
    conn = get_db()
    ensure_wa_send_log(conn)
    rows = conn.execute("""
        SELECT t.id, t.no_rekening, t.total_tagihan, t.angsuran_per_bulan,
               t.tunggakan_pokok, t.tunggakan_margin,
               n.nama, n.no_hp, n.tanggal_jt, n.marketing_nama
        FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
        WHERE t.bulan=? AND t.status='BELUM'
          AND t.total_tagihan >= 1
          AND t.kolektibilitas IN (1,2)
          AND (n.is_reschedule=0 OR n.is_reschedule IS NULL)
          AND NOT (COALESCE(t.plafond_pokok,0) > 0 AND COALESCE(t.angsuran_per_bulan,0) >= t.plafond_pokok)
          AND CAST(SUBSTR(n.tanggal_jt, 9, 2) AS INTEGER) BETWEEN ? AND ?
          AND t.no_rekening NOT IN (
              SELECT no_rekening FROM wa_send_log WHERE bulan=? AND status='terkirim')
    """, (bulan, start_day, today_day, bulan)).fetchall()
    conn.close()

    terkirim = gagal = skip = 0
    detail = []
    for row in rows:
        if not row["no_hp"]:
            skip += 1
            detail.append({"nama": row["nama"], "status": "skip", "alasan": "no HP"})
            continue
        tgl = format_tgl_jt(row["tanggal_jt"])
        angs = row["angsuran_per_bulan"] or 0
        total = row["total_tagihan"] or 0
        # Formula baru: total_tagihan = tunggakan_pokok + tunggakan_margin
        # <= 1 angsuran = tagihan bulan ini (lancar), > 1 angsuran = ada tunggakan
        if angs > 0 and total > angs + max(1000, angs * 0.10):
            actual_tung = round(total - angs)
            nominal = angs
        else:
            actual_tung = 0
            nominal = round(total)
        pesan = pesan_tagihan(row["nama"], nominal, tgl, row["marketing_nama"],
                              no_akad=row["no_rekening"], tunggakan=actual_tung)
        result = kirim_wa(row["no_hp"], pesan)
        if result.get("status") == "success":
            terkirim += 1
            detail.append({"nama": row["nama"], "status": "ok"})
            try:
                conn_wl = get_db()
                ensure_wa_send_log(conn_wl)
                conn_wl.execute(
                    "INSERT INTO wa_send_log (no_rekening, nama, no_hp, bulan, tipe, status, dikirim_oleh) VALUES (?,?,?,?,?,?,?)",
                    (row["no_rekening"], row["nama"], row["no_hp"], bulan, "blast_jt", "terkirim", session.get("nama", "admin"))
                )
                conn_wl.commit()
                conn_wl.close()
            except: pass
        else:
            gagal += 1
            detail.append({"nama": row["nama"], "status": "gagal", "alasan": str(result)})
        if len(rows) > 1:
            import time as _t
            _t.sleep(max(1, int(get_setting('delay_blast_detik', '10'))))

    # Log blast JT hari ini
    conn_log = get_db()
    ensure_blast_log(conn_log)
    conn_log.execute(
        "INSERT INTO blast_log (tipe, bulan, dilakukan_oleh, terkirim, gagal, skip, catatan) VALUES (?,?,?,?,?,?,?)",
        ("blast_jt", bulan, session.get("username", "admin"), terkirim, gagal, skip, "Blast JT Hari Ini")
    )
    conn_log.commit()
    conn_log.close()
    return jsonify({"success": True, "terkirim": terkirim,
                    "gagal": gagal, "skip": skip, "detail": detail,
                    "sync": _sync_msg(sync_info)})

@app.route("/api/bayar", methods=["POST"])
@login_required
def bayar():
    data       = request.json
    tagihan_id = data.get("tagihan_id")
    jumlah     = data.get("jumlah")
    cara_bayar = data.get("cara_bayar", "TUNAI")
    catatan    = data.get("catatan", "")

    if not tagihan_id or not jumlah:
        return jsonify({"error": "Data tidak lengkap"}), 400

    conn = get_db()
    row = conn.execute("""
        SELECT t.*, n.nama, n.no_hp, n.marketing_nama, n.marketing_id
        FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
        WHERE t.id=?
    """, (tagihan_id,)).fetchone()

    if not row:
        conn.close()
        return jsonify({"error": "Tagihan tidak ditemukan"}), 404

    if row["status"] == "LUNAS":
        conn.close()
        return jsonify({"error": "Tagihan sudah lunas"}), 400

    conn.execute(
        "UPDATE tagihan SET status='LUNAS', cara_bayar=? WHERE id=?",
        (cara_bayar, tagihan_id)
    )
    conn.execute("""
        INSERT INTO pembayaran
          (no_rekening, tagihan_id, jumlah, cara_bayar, marketing_id, catatan, dicatat_oleh)
        VALUES (?,?,?,?,?,?,?)
    """, (
        row["no_rekening"], tagihan_id, jumlah, cara_bayar,
        session.get("marketing_id"), catatan, session.get("nama")
    ))
    conn.commit()
    conn.close()

    # Kirim WA notif lunas (Dinonaktifkan)
    # if row["no_hp"]:
    #     pesan = pesan_lunas(row["nama"], jumlah, row["marketing_nama"])
    #     kirim_wa(row["no_hp"], pesan)

    return jsonify({
        "success": True,
        "message": f"Pembayaran {row['nama']} berhasil dicatat"
    })

# â”€â”€ REMINDER WA INDIVIDUAL â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/reminder/<int:tagihan_id>", methods=["POST"])
@login_required
def kirim_reminder(tagihan_id):
    data = request.json or {}
    nominal_baru = data.get("nominal")
    no_hp_baru = data.get("no_hp")
    force = data.get("force", False)

    conn = get_db()
    tag_row = conn.execute("SELECT no_rekening FROM tagihan WHERE id=?", (tagihan_id,)).fetchone()
    
    if tag_row:
        if no_hp_baru:
            import re
            clean_hp = re.sub(r'[^0-9]', '', str(no_hp_baru))
            if clean_hp.startswith("0"):
                clean_hp = "62" + clean_hp[1:]
            conn.execute("UPDATE nasabah SET no_hp=? WHERE no_rekening=?", (clean_hp, tag_row["no_rekening"]))
            
        if nominal_baru is not None:
            try:
                nominal_baru = float(nominal_baru)
                conn.execute("UPDATE tagihan SET total_tagihan=? WHERE id=?", (nominal_baru, tagihan_id))
            except:
                pass
        conn.commit()

    row = conn.execute("""
        SELECT t.*, n.nama, n.no_hp, n.marketing_nama, n.tanggal_jt, n.is_reschedule
        FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
        WHERE t.id=?
    """, (tagihan_id,)).fetchone()
    conn.close()

    if not row:
        return jsonify({"error": "Tagihan tidak ditemukan"}), 404
    if row["is_reschedule"] == 1:
        return jsonify({"error": "Nasabah reschedule — WA tagihan tidak dikirim. Hubungi langsung."}), 403
    if row["plafond_pokok"] and row["angsuran_per_bulan"] and row["angsuran_per_bulan"] >= row["plafond_pokok"]:
        return jsonify({"error": "Pinjaman bullet/musiman — tidak ditagih via WA rutin. Tangani saat jatuh tempo."}), 403
    if not row["no_hp"]:
        return jsonify({"error": "No HP nasabah belum diisi"}), 400

    if not force:
        conn_chk = get_db()
        ensure_wa_send_log(conn_chk)
        prev = conn_chk.execute(
            "SELECT COUNT(*) as c FROM wa_send_log WHERE no_rekening=? AND bulan=? AND status='terkirim'",
            (row["no_rekening"], row["bulan"])
        ).fetchone()
        conn_chk.close()
        if prev and prev["c"] > 0:
            return jsonify({"warning": True, "message": "Nasabah ini sudah pernah dikirim WA " + str(prev["c"]) + "x bulan ini. Kirim lagi?", "count": prev["c"]})

    tgl = format_tgl_jt(row["tanggal_jt"])
    angs_pb = row["angsuran_per_bulan"] or 0
    total_th = row["total_tagihan"] or 0

    # Jika tagihan = 0, tidak perlu kirim WA
    if total_th < 1:
        return jsonify({"error": "Tagihan Rp 0 — tidak perlu kirim WA"}), 400

    # Formula baru: total > 1 angsuran = ada tunggakan lama
    if angs_pb > 0 and total_th > angs_pb + max(1000, angs_pb * 0.10):
        actual_tung = round(total_th - angs_pb)
        nominal_pesan = angs_pb
    else:
        actual_tung = 0
        nominal_pesan = round(total_th)
    pesan = pesan_tagihan(row["nama"], nominal_pesan, tgl, row["marketing_nama"], no_akad=row["no_rekening"], tunggakan=actual_tung)
    result = kirim_wa(row["no_hp"], pesan)
    wa_status = "terkirim" if (result and result.get("status") != "skip") else "gagal"
    conn2 = get_db()
    ensure_wa_send_log(conn2)
    conn2.execute(
        "INSERT INTO wa_send_log (no_rekening, nama, no_hp, bulan, tipe, status, dikirim_oleh) VALUES (?,?,?,?,?,?,?)",
        (row["no_rekening"], row["nama"], row["no_hp"], row["bulan"], "individual", wa_status, session.get("nama", "admin"))
    )
    conn2.commit()
    conn2.close()
    return jsonify({"success": True, "wa_result": result})

# â”€â”€ BLAST REMINDER â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/reminder/preview_blast", methods=["POST"])
@admin_required
def preview_blast():
    data = request.json
    bulan = data.get("bulan", datetime.now().strftime("%Y-%m"))
    hanya_hari_ini = data.get("hanya_hari_ini", False)

    conn = get_db()
    where = "WHERE t.bulan=? AND t.status='BELUM' AND n.no_hp IS NOT NULL AND n.no_hp != '' AND t.kolektibilitas IN (1,2) AND (n.is_reschedule=0 OR n.is_reschedule IS NULL) AND NOT (COALESCE(t.plafond_pokok,0) > 0 AND COALESCE(t.angsuran_per_bulan,0) >= t.plafond_pokok)"
    params = [bulan]

    rows = conn.execute(f"""
        SELECT t.id, t.total_tagihan, t.kolektibilitas, n.nama, n.no_rekening, n.no_hp, n.marketing_nama, n.tanggal_jt, n.is_reschedule
        FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
        {where}
        ORDER BY t.kolektibilitas DESC, n.tanggal_jt ASC
    """, params).fetchall()
    conn.close()

    lancar = []
    bermasalah = []
    hari_ini_str = datetime.now().strftime("%d")
    hari_ini_str_alt = hari_ini_str[1:] if hari_ini_str.startswith("0") else hari_ini_str

    for r in rows:
        d = dict(r)
        tgl = format_tgl_jt(d["tanggal_jt"])
        if hanya_hari_ini and tgl not in (hari_ini_str, hari_ini_str_alt):
            continue
            
        if d["kolektibilitas"] == 1:
            lancar.append(d)
        else:
            bermasalah.append(d)

    return jsonify({"success": True, "lancar": lancar, "bermasalah": bermasalah})


@app.route("/api/blast/task/<task_id>")
@login_required
def get_blast_task(task_id):
    """Cek status background blast task"""
    conn = get_db()
    row = conn.execute(
        "SELECT * FROM blast_tasks WHERE task_id=?", [task_id]
    ).fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "Task tidak ditemukan"}), 404
    return jsonify(dict(row))

@app.route("/api/reminder/execute_blast", methods=["POST"])
@admin_required
def execute_blast():
    import time
    data = request.json
    bulan = data.get("bulan", datetime.now().strftime("%Y-%m"))
    hanya_hari_ini = data.get("hanya_hari_ini", False)
    updates = data.get("updates", [])

    # Auto-sync MESPro dulu, SEBELUM apply override manual (agar override admin menang).
    # Tetap lanjut walau sync gagal -> blast pakai data terakhir.
    sync_info = _run_sync_mespro()

    conn = get_db()
    for up in updates:
        tag_id = up.get("id")
        nom = up.get("nominal")
        if tag_id and nom is not None:
            conn.execute("UPDATE tagihan SET total_tagihan=? WHERE id=?", (float(nom), tag_id))
    conn.commit()

    where = "WHERE t.bulan=? AND t.status='BELUM' AND n.no_hp IS NOT NULL AND n.no_hp != '' AND t.kolektibilitas IN (1,2) AND (n.is_reschedule=0 OR n.is_reschedule IS NULL) AND NOT (COALESCE(t.plafond_pokok,0) > 0 AND COALESCE(t.angsuran_per_bulan,0) >= t.plafond_pokok)"
    params = [bulan]

    rows = [dict(r) for r in conn.execute(f"""
        SELECT t.id, t.no_rekening, t.total_tagihan, t.tunggakan_pokok, t.tunggakan_margin,
               t.angsuran_per_bulan,
               n.nama, n.no_hp, n.marketing_nama, n.tanggal_jt
        FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
        {where}
    """, params).fetchall()]
    conn.close()

    import uuid as _uuid
    import threading as _threading

    task_id = _uuid.uuid4().hex[:12]
    username = session.get("username", "admin")
    total_rows = len(rows)
    catatan_blast = "Blast Hari Ini" if hanya_hari_ini else "Blast Semua"

    # Buat record task di DB
    conn_task = get_db()
    ensure_blast_log(conn_task)
    conn_task.execute(
        "INSERT INTO blast_tasks (task_id, tipe, bulan, status, terkirim, gagal, total, dilakukan_oleh, catatan) VALUES (?,?,?,?,?,?,?,?,?)",
        (task_id, "execute_blast", bulan, "running", 0, 0, total_rows, username, catatan_blast)
    )
    conn_task.commit()
    conn_task.close()

    def _do_blast_bg(rows_copy, bulan_copy, hanya_hari_ini_copy, task_id_copy, username_copy, catatan_copy):
        terkirim = gagal = 0
        hari_ini_str = datetime.now().strftime("%d")
        hari_ini_str_alt = hari_ini_str[1:] if hari_ini_str.startswith("0") else hari_ini_str
        try:
            for row in rows_copy:
                tgl = format_tgl_jt(row["tanggal_jt"])
                if hanya_hari_ini_copy and tgl not in (hari_ini_str, hari_ini_str_alt):
                    continue

                angs_pb = row["angsuran_per_bulan"] or 0
                total_th = row["total_tagihan"] or 0
                # Formula baru: total > 1 angsuran = ada tunggakan lama
                if angs_pb > 0 and total_th > angs_pb + max(1000, angs_pb * 0.10):
                    actual_tung = round(total_th - angs_pb)
                    nominal_pesan = angs_pb
                else:
                    actual_tung = 0
                    nominal_pesan = round(total_th)

                pesan = pesan_tagihan(row["nama"], nominal_pesan, tgl, row["marketing_nama"],
                                      no_akad=row["no_rekening"], tunggakan=actual_tung)
                result = kirim_wa(row["no_hp"], pesan)
                if result.get("status") == True:
                    terkirim += 1
                    try:
                        import sqlite3 as _sq
                        _cl = _sq.connect(os.path.join(os.path.dirname(__file__), "data", "koperasi.db"))
                        _cl.execute(
                            "INSERT INTO wa_send_log (no_rekening, nama, no_hp, bulan, tipe, status, dikirim_oleh) VALUES (?,?,?,?,?,?,?)",
                            (row["no_rekening"], row["nama"], row["no_hp"], bulan_copy, "blast", "terkirim", username_copy)
                        )
                        _cl.commit()
                        _cl.close()
                    except: pass
                else:
                    gagal += 1

                # Update progress di DB
                try:
                    c2 = get_db()
                    c2.execute(
                        "UPDATE blast_tasks SET terkirim=?, gagal=? WHERE task_id=?",
                        (terkirim, gagal, task_id_copy)
                    )
                    c2.commit()
                    c2.close()
                except Exception:
                    pass

                time.sleep(max(1, int(get_setting("delay_blast_detik", "10"))))

            # Selesai - update status done + log ke blast_log
            c3 = get_db()
            ensure_blast_log(c3)
            c3.execute(
                "UPDATE blast_tasks SET status='done', selesai_at=datetime('now','localtime') WHERE task_id=?",
                (task_id_copy,)
            )
            c3.execute(
                "INSERT INTO blast_log (tipe, bulan, dilakukan_oleh, terkirim, gagal, skip, catatan) VALUES (?,?,?,?,?,?,?)",
                ("execute_blast", bulan_copy, username_copy, terkirim, gagal, 0, catatan_copy)
            )
            c3.commit()
            c3.close()
        except Exception as e:
            try:
                c_err = get_db()
                c_err.execute(
                    "UPDATE blast_tasks SET status='error', catatan=? WHERE task_id=?",
                    (str(e)[:200], task_id_copy)
                )
                c_err.commit()
                c_err.close()
            except Exception:
                pass

    t = _threading.Thread(
        target=_do_blast_bg,
        args=(rows, bulan, hanya_hari_ini, task_id, username, catatan_blast),
        daemon=True
    )
    t.start()
    return jsonify({"success": True, "task_id": task_id, "total": total_rows,
                    "background": True, "sync": _sync_msg(sync_info)})

# â”€â”€ UPDATE NO HP â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/nasabah/<no_rek>/hp", methods=["PUT"])
@login_required
def update_hp(no_rek):
    data  = request.json
    no_hp = data.get("no_hp", "").strip()
    conn  = get_db()
    conn.execute("UPDATE nasabah SET no_hp=? WHERE no_rekening=?", (no_hp, no_rek))
    conn.commit()
    conn.close()
    return jsonify({"success": True})

# â”€â”€ HISTORI PEMBAYARAN â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/histori")
@login_required
def histori():
    conn         = get_db()
    marketing_id = session.get("marketing_id")
    role         = get_user_role()

    mkt_cond = ""
    params   = []
    if role not in ("admin","leader","petugas") and marketing_id:
        mkt_cond = "AND n.marketing_id=?"
        params.append(marketing_id)

    # Gabung pembayaran manual + tagihan LUNAS dari import
    rows = conn.execute(f"""
        SELECT n.nama, p.no_rekening, p.cara_bayar,
               p.dicatat_oleh, p.tanggal, p.jumlah, p.catatan
        FROM pembayaran p JOIN nasabah n ON p.no_rekening = n.no_rekening
        WHERE 1=1 {mkt_cond}
        UNION ALL
        SELECT n.nama, t.no_rekening,
               COALESCE(t.cara_bayar,'TUNAI') as cara_bayar,
               'Import' as dicatat_oleh,
               CASE WHEN length(CAST(t.tgl_bayar AS TEXT))=8
                    THEN substr(CAST(t.tgl_bayar AS TEXT),1,4)||'-'||
                         substr(CAST(t.tgl_bayar AS TEXT),5,2)||'-'||
                         substr(CAST(t.tgl_bayar AS TEXT),7,2)
                    ELSE CAST(t.tgl_bayar AS TEXT) END as tanggal,
               t.angsuran_per_bulan as jumlah,
               t.bulan as catatan
        FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
        WHERE (t.status='LUNAS' OR t.total_tagihan < 1)
        {mkt_cond}
        ORDER BY tanggal DESC
        LIMIT 300
    """, params + params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

# â”€â”€ IMPORT EXCEL â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

# ── AUTO IMPORT (token-based, tanpa session) ─────────────────────────────────
@app.route("/api/import/auto", methods=["POST"])
def import_excel_auto():
    """Endpoint untuk upload otomatis via curl dari Windows (pakai API token)."""
    # Validasi token
    token_file = os.path.join(os.path.dirname(__file__), 'data', 'api_token.txt')
    try:
        with open(token_file) as _tf:
            valid_token = _tf.read().strip()
    except Exception:
        return jsonify({"error": "Server belum dikonfigurasi token"}), 500

    req_token = request.headers.get('X-BMT-Token', '').strip()
    if not req_token or not hmac.compare_digest(req_token, valid_token):
        return jsonify({"error": "Token tidak valid"}), 403

    if "file" not in request.files:
        return jsonify({"error": "File tidak ditemukan"}), 400

    f        = request.files["file"]
    safe_fn  = secure_filename(f.filename) or "upload.xlsx"
    filepath = os.path.join(os.path.dirname(__file__), "data", f"upload_{safe_fn}")
    f.save(filepath)

    from init_db import import_excel
    result = import_excel(filepath, diimport_oleh="[AUTO-UPLOAD]")
    try:
        os.remove(filepath)
    except:
        pass
    return jsonify(result)

@app.route("/api/import", methods=["POST"])
@admin_required
def import_excel_route():
    if "file" not in request.files:
        return jsonify({"error": "File tidak ditemukan"}), 400

    f        = request.files["file"]
    safe_fn  = secure_filename(f.filename) or "upload.xlsx"
    filepath = os.path.join(os.path.dirname(__file__), "data", f"upload_{safe_fn}")
    f.save(filepath)

    from init_db import import_excel
    result = import_excel(filepath, diimport_oleh=session.get("nama", "admin"))
    try:
        os.remove(filepath)
    except:
        pass
    return jsonify(result)

# â”€â”€ HISTORI IMPORT â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

# -- SYNC MESPRO --------------------------------------------------
@app.route("/api/sync/mespro", methods=["POST"])
@admin_required
def sync_mespro_endpoint():
    """Trigger sync manual dari MESPro MySQL ke koperasi.db"""
    return jsonify(_run_sync_mespro())

@app.route("/api/sync/status")
@admin_required
def sync_status():
    """Status sync MESPro terakhir (dari data/sync_state.json yang ditulis
    sync_mespro.py). Dipakai admin untuk lihat kapan sync terakhir berjalan
    & apakah berhasil -- termasuk run cron 10 menit yang tak tercatat di
    import_log."""
    import json as _json
    path = os.path.join(os.path.dirname(__file__), "data", "sync_state.json")
    if not os.path.exists(path):
        return jsonify({"ada": False})
    try:
        st = _json.load(open(path))
        st.pop("last_notif_ts", None)  # internal, tak perlu dikirim ke UI
        st["ada"] = True
        return jsonify(st)
    except Exception as e:
        return jsonify({"ada": False, "error": str(e)})

@app.route("/api/import/log")
@admin_required
def import_log():
    conn = get_db()
    rows = conn.execute("""
        SELECT bulan, filepath, nasabah_baru, nasabah_update, nasabah_nonaktif,
               tagihan_baru, tagihan_update, diimport_oleh, waktu
        FROM import_log ORDER BY waktu DESC LIMIT 20
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])
    # ── TEMPLATE PESAN ─────────────────────────────────────────────
@app.route("/api/template", methods=["GET"])
@admin_required
def get_template():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS template_pesan (
            id TEXT PRIMARY KEY,
            judul TEXT,
            isi TEXT
        )
    """)
    # Default template kalau belum ada
    tpl_tagihan = """Bismillaahirrahmaanirrahiim
Assalamu 'alaikum warahmatullaahi wabarakaatuh

Kepada
Ykh. Bp/ibu {nasabah_nama}
Di tempat
Sholawat dan salam semoga tetap tercurahkan kepada nabi Muhammad saw, sahabat dan generasi penerusnya. aamiin
Mohon maaf, mengingatkan bahwa angsuran bpk/ibu di bulan ini sudah jatuh tempo pada tanggal {jatuh_tempo} dengan No akad : {no_akad} , sebesar : {total}

Pembayaran bisa dilakukan dengan :
1.	Bayar tunai di kantor
2.	Potong tabungan
3.	Transfer ke bank :
✅ BSI no rek. 7778880231 an. KSPPS BMT AMAL MUSLIM
✅ BANK JATENG SYARIAH no rek. 6133007001 an. KSPPS BMT AMAL MUSLIM
Demikian semoga Alloh memberikan kemudahan kepada kita semua, aamiin

Wassalamu 'alaikum wr wb.

Hormat kami,
Admin
BMT Amal Muslim

) Abaikan apabila angsuran sudah terbayar"""

    tpl_tagihan_tunggakan = """Bismillaahirrahmaanirrahiim
Assalamu 'alaikum warahmatullaahi wabarakaatuh

Kepada
Ykh. Bp/ibu {nasabah_nama}
Di tempat
Sholawat dan salam semoga tetap tercurahkan kepada nabi Muhammad saw, sahabat dan generasi penerusnya. aamiin
Mohon maaf, mengingatkan bahwa angsuran bpk/ibu di bulan ini sudah jatuh tempo tanggal {jatuh_tempo} dengan No akad : {no_akad} , sebesar : {total} dan tunggakan angsuran bulan sebelumnya {tunggakan} Jumlah Total {total_keseluruhan}

Pembayaran bisa dilakukan dengan :
1.	Bayar tunai di kantor
2.	Potong tabungan
3.	Transfer ke bank :
✅ BSI no rek. 7778880231 an. KSPPS BMT AMAL MUSLIM
✅ BANK JATENG SYARIAH no rek. 6133007001 an. KSPPS BMT AMAL MUSLIM

Demikian semoga Alloh memberikan kemudahan kepada kita semua, aamiin

Wassalamu 'alaikum wr wb.

Hormat kami,
Admin
BMT Amal Muslim

) *Abaikan apabila angsuran sudah terbayar"""

    defaults = [
        ("tagihan", "Reminder Tagihan", tpl_tagihan),
        ("tagihan_tunggakan", "Reminder Tagihan (Ada Tunggakan)", tpl_tagihan_tunggakan),
        ("lunas", "Konfirmasi Lunas", """Assalamu'alaikum, {nasabah_nama} 🙏

✅ Pembayaran Anda telah *berhasil dicatat*!

💰 *Jumlah Bayar:* {jumlah}
📅 *Tanggal:* {tgl_sekarang}
👤 *Marketing:* {marketing_nama}

Terima kasih atas kepercayaan Anda 🙏
*KSPPS BMT Amal Muslim Wonogiri*"""),
    ]
    for id, judul, isi in defaults:
        conn.execute("INSERT OR IGNORE INTO template_pesan VALUES (?,?,?)", (id, judul, isi))
    # Update tagihan templates in case they were already saved with old format
    conn.execute("UPDATE template_pesan SET isi=? WHERE id='tagihan'", (tpl_tagihan,))
    conn.execute("UPDATE template_pesan SET judul=?, isi=? WHERE id='tagihan_tunggakan'",
                 ("Reminder Tagihan (Ada Tunggakan)", tpl_tagihan_tunggakan))
    conn.commit()
    rows = conn.execute("SELECT * FROM template_pesan").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/template/<id>", methods=["PUT"])
@admin_required
def update_template(id):
    data = request.json
    isi = data.get("isi", "").strip()
    if not isi:
        return jsonify({"error": "Isi pesan tidak boleh kosong"}), 400
    conn = get_db()
    conn.execute("UPDATE template_pesan SET isi=? WHERE id=?", (isi, id))
    conn.commit()
    conn.close()
    return jsonify({"success": True})

# â”€â”€ SERVE FRONTEND â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/")
@app.route("/<path:path>")
def index(path=""):
    return render_template("index.html")

# â”€â”€ STARTUP: init DB kalau belum ada â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def startup():
    if not os.path.exists(DB_PATH):
        print("\U0001f527 Database belum ada, menjalankan init_db...")
        from init_db import init_db
        init_db()
    else:
        print("\u2705 Database ditemukan, skip init.")
    # Pastikan tabel blast_log/blast_tasks ada, dan reset stuck tasks
    try:
        _sc = get_db()
        ensure_blast_log(_sc)
        ensure_wa_send_log(_sc)
        _sc.execute("UPDATE blast_tasks SET status='interrupted', catatan='App restart' WHERE status='running'")
        _sc.commit()
        _sc.close()
    except Exception:
        pass
    # Init tabel modul keanggotaan (idempotent)
    try:
        from init_keanggotaan import init_keanggotaan
        init_keanggotaan(DB_PATH, verbose=False)
    except Exception as _e:
        print(f"[WARN] init_keanggotaan: {_e}")

startup()

# ── RESCHEDULE ─────────────────────────────────────────────────────────────
@app.route("/api/reschedule/pending", methods=["GET"])
@login_required
def get_pending_reschedule():
    conn = get_db()
    rows = conn.execute("SELECT no_rekening, nama FROM nasabah WHERE is_reschedule = -1").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/reschedule/confirm", methods=["POST"])
@login_required
def confirm_reschedule():
    data = request.json
    no_rek = data.get("no_rekening")
    is_reschedule = 1 if data.get("is_reschedule") else 0
    conn = get_db()
    conn.execute("UPDATE nasabah SET is_reschedule=? WHERE no_rekening=?", (is_reschedule, no_rek))
    conn.commit()
    conn.close()
    return jsonify({"success": True})

# ── AUTO MIGRATION ───────────────────────────────────────────────────────
def check_db_schema():
    conn = get_db()
    c = conn.cursor()
    c.execute("PRAGMA table_info(nasabah)")
    cols = [r["name"] for r in c.fetchall()]
    if "tgl_realisasi" not in cols:
        try: c.execute("ALTER TABLE nasabah ADD COLUMN tgl_realisasi TEXT")
        except: pass
    if "is_reschedule" not in cols:
        try: c.execute("ALTER TABLE nasabah ADD COLUMN is_reschedule INTEGER DEFAULT 0")
        except: pass
    conn.commit()
    conn.close()

check_db_schema()

# ── DASHBOARD MARKETING REAL-TIME ──────────────────────────────

@app.route("/api/dashboard/tren-tahunan")
@login_required
def dashboard_tren_tahunan():
    tahun = request.args.get("tahun", datetime.now().strftime("%Y"))
    conn = get_db()
    role = get_user_role(); marketing_id = session.get("marketing_id")
    is_mkt = role not in ("admin", "leader", "petugas") and marketing_id
    mkt_join = " JOIN nasabah n ON t.no_rekening=n.no_rekening" if is_mkt else ""
    mkt_cond = " AND n.marketing_id=?" if is_mkt else ""
    BULAN_LABEL = ["","Jan","Feb","Mar","Apr","Mei","Jun","Jul","Agu","Sep","Okt","Nov","Des"]

    bulanan = []
    for m in range(1, 13):
        bulan = f"{tahun}-{m:02d}"
        row = conn.execute(f"""
            SELECT COUNT(*) as total,
                SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN 1 ELSE 0 END) as lunas,
                SUM(CASE WHEN (t.status='BELUM') THEN 1 ELSE 0 END) as belum,
                SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN t.angsuran_per_bulan ELSE 0 END) as terkumpul,
                SUM(CASE WHEN (t.status='BELUM' AND t.total_tagihan >= 1) THEN t.total_tagihan ELSE 0 END) as tunggakan
            FROM tagihan t{mkt_join} WHERE t.bulan=?{mkt_cond}
        """, [bulan] + ([marketing_id] if is_mkt else [])).fetchone()
        bulanan.append({
            "bulan": bulan, "label": BULAN_LABEL[m],
            "total": row[0] or 0, "lunas": row[1] or 0, "belum": row[2] or 0,
            "terkumpul": row[3] or 0, "tunggakan": row[4] or 0,
        })

    if is_mkt:
        mkt_rows = conn.execute(
            "SELECT DISTINCT marketing_nama FROM nasabah WHERE marketing_id=? AND marketing_nama IS NOT NULL",
            [marketing_id]).fetchall()
    else:
        mkt_rows = conn.execute(
            "SELECT DISTINCT marketing_nama FROM nasabah WHERE aktif=1 AND marketing_nama IS NOT NULL ORDER BY marketing_nama"
        ).fetchall()

    per_marketing = []
    for mkt_row in mkt_rows:
        mkt = mkt_row[0]
        mkt_data = []
        for m in range(1, 13):
            bulan = f"{tahun}-{m:02d}"
            r = conn.execute("""
                SELECT COUNT(*) as total,
                    SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN 1 ELSE 0 END) as lunas,
                    SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN t.angsuran_per_bulan ELSE 0 END) as terkumpul
                FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
                WHERE t.bulan=? AND n.marketing_nama=?
            """, [bulan, mkt]).fetchone()
            mkt_data.append({"bulan": bulan, "total": r[0] or 0, "lunas": r[1] or 0, "terkumpul": r[2] or 0})
        per_marketing.append({"nama": mkt, "data": mkt_data})

    conn.close()
    return jsonify({"tahun": tahun, "bulanan": bulanan, "per_marketing": per_marketing})

@app.route("/api/dashboard/marketing")
@login_required
def dashboard_marketing():
    conn = get_db()
    bulan = request.args.get("bulan", datetime.now().strftime("%Y-%m"))
    role = get_user_role()
    marketing_id = session.get("marketing_id")

    if role in ("admin","leader","petugas"):
        rows = conn.execute("""
            SELECT n.marketing_nama, n.marketing_id,
                COUNT(*) as total,
                SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN 1 ELSE 0 END) as lunas,
                SUM(CASE WHEN (t.status='BELUM') THEN 1 ELSE 0 END) as belum,
                SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN t.angsuran_per_bulan ELSE 0 END) as nominal_lunas,
                SUM(CASE WHEN (t.status='BELUM' AND t.total_tagihan >= 1) THEN t.total_tagihan ELSE 0 END) as nominal_belum,
                SUM(t.angsuran_per_bulan) as total_tagihan
            FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
            WHERE t.bulan=?
              AND n.marketing_nama NOT GLOB '[0-9]*'
            GROUP BY n.marketing_nama, n.marketing_id ORDER BY lunas DESC
        """, [bulan]).fetchall()
    else:
        rows = conn.execute("""
            SELECT n.marketing_nama, n.marketing_id,
                COUNT(*) as total,
                SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN 1 ELSE 0 END) as lunas,
                SUM(CASE WHEN (t.status='BELUM') THEN 1 ELSE 0 END) as belum,
                SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN t.angsuran_per_bulan ELSE 0 END) as nominal_lunas,
                SUM(CASE WHEN (t.status='BELUM' AND t.total_tagihan >= 1) THEN t.total_tagihan ELSE 0 END) as nominal_belum,
                SUM(t.angsuran_per_bulan) as total_tagihan
            FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
            WHERE t.bulan=? AND n.marketing_id=?
            GROUP BY n.marketing_nama, n.marketing_id
        """, [bulan, marketing_id]).fetchall()

    bulan_num = bulan.replace('-', '')  # e.g. '202606'
    is_marketing = role not in ("admin", "leader", "petugas")
    mkt_join_filter = "AND n.marketing_id=?" if is_marketing else ""
    mkt_args_bn = [bulan, bulan_num] + ([marketing_id] if is_marketing else [])
    mkt_args_b  = [bulan] + ([marketing_id] if is_marketing else [])

    # FIX: tgl_bayar punya 2 format (YYYYMMDD & YYYY-MM-DD). Normalisasi dgn replace('-').
    # Hitung semua pembayaran yg tgl_bayar-nya di bulan ini (apapun status/format).
    tren = conn.execute("""
        SELECT substr(replace(t.tgl_bayar,'-',''), 7, 2) as hari,
            COUNT(*) as jumlah_transaksi,
            SUM(t.angsuran_per_bulan) as total_nominal
        FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
        WHERE t.bulan=?
          AND t.tgl_bayar IS NOT NULL AND t.tgl_bayar != ''
          AND substr(replace(t.tgl_bayar,'-',''), 1, 6) = ?
          AND substr(replace(t.tgl_bayar,'-',''), 1, 8) <= strftime('%Y%m%d', 'now', 'localtime')
          """ + mkt_join_filter + """
        GROUP BY hari ORDER BY hari ASC
    """, mkt_args_bn).fetchall()

    kolek = conn.execute("""
        SELECT t.kolektibilitas, COUNT(*) as total,
            SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN 1 ELSE 0 END) as lunas,
            SUM(t.saldo_pinjaman) as nominal
        FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
        WHERE t.bulan=? """ + mkt_join_filter + """
        GROUP BY t.kolektibilitas ORDER BY t.kolektibilitas ASC
    """, mkt_args_b).fetchall()

    top_tunggak = conn.execute("""
        SELECT n.nama, n.no_rekening, n.marketing_nama, t.total_tagihan,
               t.kolektibilitas, t.saldo_pinjaman
        FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
        WHERE t.bulan=? AND t.status='BELUM' """ + mkt_join_filter + """
        ORDER BY t.saldo_pinjaman DESC LIMIT 25
    """, mkt_args_b).fetchall()

    npf_row = conn.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN t.kolektibilitas >= 3 THEN t.saldo_pinjaman ELSE 0 END), 0) as npf_nominal,
            COALESCE(SUM(t.saldo_pinjaman), 0) as total_saldo,
            COUNT(CASE WHEN t.kolektibilitas >= 3 THEN 1 END) as npf_count,
            COUNT(*) as total_count
        FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
        WHERE t.bulan=? """ + mkt_join_filter + """
    """, mkt_args_b).fetchone()
    npf_pct = round(npf_row[0] / npf_row[1] * 100, 2) if npf_row[1] > 0 else 0

    conn.close()
    return jsonify({
        "rekap_marketing": [dict(r) for r in rows],
        "tren_harian": [dict(r) for r in tren],
        "kolektibilitas": [dict(r) for r in kolek],
        "top_tunggak": [dict(r) for r in top_tunggak],
        "npf": {
            "pct": npf_pct,
            "nominal": npf_row[0],
            "total_saldo": npf_row[1],
            "npf_count": npf_row[2],
            "total_count": npf_row[3]
        },
        "bulan": bulan
    })

@app.route("/api/nasabah/<no_rek>/riwayat")
@login_required
def riwayat_anggota(no_rek):
    conn = get_db()
    nasabah = conn.execute("SELECT * FROM nasabah WHERE no_rekening=?", (no_rek,)).fetchone()
    if not nasabah:
        conn.close()
        return jsonify({"error": "Nasabah tidak ditemukan"}), 404
    tagihan_list = conn.execute("""
        SELECT t.*, p.jumlah as jumlah_bayar, p.tanggal as tgl_bayar_app,
               p.cara_bayar as cara_bayar_app, p.dicatat_oleh
        FROM tagihan t LEFT JOIN pembayaran p ON t.id = p.tagihan_id
        WHERE t.no_rekening=? ORDER BY t.bulan DESC
    """, (no_rek,)).fetchall()
    conn.close()
    return jsonify({"nasabah": dict(nasabah), "tagihan": [dict(r) for r in tagihan_list]})

@app.route("/api/dashboard/belum-minggu-ini")
@login_required
def belum_minggu_ini():
    conn = get_db()
    bulan = request.args.get("bulan", datetime.now().strftime("%Y-%m"))
    role = get_user_role()
    marketing_id = session.get("marketing_id")
    today = datetime.now()
    start_week = today - timedelta(days=today.weekday())
    end_week = start_week + timedelta(days=6)
    rows = conn.execute("""
        SELECT n.nama, n.no_rekening, n.no_hp, n.marketing_nama,
               n.tanggal_jt, t.total_tagihan, t.kolektibilitas, t.id as tagihan_id,
               n.is_reschedule
        FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
        WHERE t.bulan=? AND t.status='BELUM'
        """ + ("" if role in ("admin","leader","petugas") else "AND n.marketing_id=?") + """
        ORDER BY t.kolektibilitas DESC, t.total_tagihan DESC
    """, [bulan] if role in ("admin","leader","petugas") else [bulan, marketing_id]).fetchall()
    hasil = []
    for r in rows:
        d = dict(r)
        tgl_raw = str(d.get("tanggal_jt") or "")
        try:
            tgl_num = int(''.join(filter(str.isdigit, tgl_raw[:2])) or 0)
            if start_week.day <= tgl_num <= end_week.day:
                d["tgl_jt_num"] = tgl_num
                hasil.append(d)
        except:
            pass
    conn.close()
    return jsonify({"data": hasil, "total": len(hasil)})

@app.route("/api/dashboard/ranking")
@login_required
def ranking_marketing():
    conn = get_db()
    bulan = request.args.get("bulan", datetime.now().strftime("%Y-%m"))
    role = get_user_role(); marketing_id = session.get("marketing_id")
    is_mkt = role not in ("admin", "leader", "petugas") and marketing_id
    mkt_cond = " AND n.marketing_id=?" if is_mkt else ""
    rows = conn.execute(f"""
        SELECT n.marketing_nama, COUNT(*) as total_nasabah,
            SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN 1 ELSE 0 END) as lunas,
            SUM(CASE WHEN (t.status='BELUM') THEN 1 ELSE 0 END) as belum,
            SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN t.angsuran_per_bulan ELSE 0 END) as nominal_lunas,
            ROUND(CAST(SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN 1 ELSE 0 END) AS FLOAT)/CAST(COUNT(*) AS FLOAT)*100,1) as pct_kolektibilitas
        FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
        WHERE t.bulan=?
          AND n.marketing_nama NOT GLOB '[0-9]*'{mkt_cond}
        GROUP BY n.marketing_nama ORDER BY pct_kolektibilitas DESC, nominal_lunas DESC
    """, [bulan] + ([marketing_id] if is_mkt else [])).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

# ── LAPORAN & EXPORT (Fase 2) ──────────────────────────────────
def _laporan_rows(conn, bulan):
    role = get_user_role(); marketing_id = session.get("marketing_id")
    is_mkt = role not in ("admin", "leader", "petugas") and marketing_id
    mkt_cond = " AND n.marketing_id=?" if is_mkt else ""
    rows = conn.execute(f"""
        SELECT n.marketing_nama,
            COUNT(*) as total,
            SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN 1 ELSE 0 END) as sudah,
            SUM(CASE WHEN (t.status='BELUM') THEN 1 ELSE 0 END) as belum,
            SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN t.angsuran_per_bulan ELSE 0 END) as terkumpul,
            SUM(CASE WHEN (t.status='BELUM' AND t.total_tagihan>=1) THEN t.total_tagihan ELSE 0 END) as tunggakan
        FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
        WHERE t.bulan=? AND n.marketing_nama NOT GLOB '[0-9]*'{mkt_cond}
        GROUP BY n.marketing_nama ORDER BY belum DESC, tunggakan DESC
    """, [bulan] + ([marketing_id] if is_mkt else [])).fetchall()
    return [dict(r) for r in rows]

@app.route("/api/laporan/rekap")
@login_required
def laporan_rekap():
    bulan = request.args.get("bulan", datetime.now().strftime("%Y-%m"))
    conn = get_db()
    data = _laporan_rows(conn, bulan)
    conn.close()
    tot = {
        "total":     sum(r["total"] for r in data),
        "sudah":     sum(r["sudah"] for r in data),
        "belum":     sum(r["belum"] for r in data),
        "terkumpul": sum(r["terkumpul"] or 0 for r in data),
        "tunggakan": sum(r["tunggakan"] or 0 for r in data),
    }
    return jsonify({"bulan": bulan, "rows": data, "total": tot})

@app.route("/api/laporan/export")
@login_required
def laporan_export():
    import csv, io as _io
    from flask import Response
    bulan = request.args.get("bulan", datetime.now().strftime("%Y-%m"))
    conn = get_db()
    data = _laporan_rows(conn, bulan)
    conn.close()
    out = _io.StringIO()
    out.write("﻿")  # BOM agar Excel baca UTF-8
    w = csv.writer(out)
    w.writerow(["Marketing", "Total Nasabah", "Sudah Bayar", "Belum Bayar",
                "Terkumpul (Rp)", "Tunggakan (Rp)", "% Kolektibilitas"])
    for r in data:
        pct = round(r["sudah"] / r["total"] * 100, 1) if r["total"] else 0
        w.writerow([r["marketing_nama"], r["total"], r["sudah"], r["belum"],
                    int(r["terkumpul"] or 0), int(r["tunggakan"] or 0), pct])
    return Response(out.getvalue(), mimetype="text/csv; charset=utf-8",
                    headers={"Content-Disposition": "attachment; filename=laporan_bmt_" + bulan + ".csv"})

@app.route("/api/jadwal-notif", methods=["GET"])
@admin_required
def get_jadwal():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS jadwal_notif (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipe TEXT NOT NULL UNIQUE, jam TEXT NOT NULL, aktif INTEGER DEFAULT 1,
            keterangan TEXT, dibuat_oleh TEXT,
            dibuat_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    defaults = [
        ("reminder_h3","08:00",1,"Reminder jatuh tempo hari ini ke anggota"),
        ("laporan_harian","17:00",1,"Laporan harian ke admin & marketing"),
        ("rekap_mingguan","17:00",1,"Rekap mingguan per marketing (Jumat)"),
    ]
    for tipe, jam, aktif, ket in defaults:
        conn.execute("INSERT OR IGNORE INTO jadwal_notif (tipe,jam,aktif,keterangan) VALUES (?,?,?,?)",(tipe,jam,aktif,ket))
    conn.commit()
    rows = conn.execute("SELECT * FROM jadwal_notif ORDER BY id").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/jadwal-notif/<int:id>", methods=["PUT"])
@admin_required
def update_jadwal(id):
    data = request.json
    conn = get_db()
    conn.execute("UPDATE jadwal_notif SET jam=?, aktif=?, dibuat_oleh=? WHERE id=?",
        (data.get("jam","08:00"), int(data.get("aktif",1)), session.get("nama"), id))
    conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route("/api/notif/laporan-harian", methods=["POST"])
@admin_required
def kirim_laporan_harian():
    conn = get_db()
    bulan = datetime.now().strftime("%Y-%m")
    today = datetime.now().strftime("%d/%m/%Y")
    stats = conn.execute("""
        SELECT COUNT(*) as total_nasabah,
            SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN 1 ELSE 0 END) as sudah_bayar,
            SUM(CASE WHEN (t.status='BELUM') THEN 1 ELSE 0 END) as belum_bayar,
            SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN t.angsuran_per_bulan ELSE 0 END) as terkumpul,
            SUM(CASE WHEN (t.status='BELUM' AND t.total_tagihan >= 1) THEN t.total_tagihan ELSE 0 END) as tunggakan
        FROM tagihan t WHERE t.bulan=?
    """, [bulan]).fetchone()
    transaksi = conn.execute("""
        SELECT COUNT(*) as jumlah, SUM(p.jumlah) as total FROM pembayaran p
        WHERE DATE(p.tanggal) = DATE('now', 'localtime')
    """).fetchone()
    marketing_rows = conn.execute("""
        SELECT n.marketing_nama, COUNT(*) as total,
            SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN 1 ELSE 0 END) as lunas,
            SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN t.angsuran_per_bulan ELSE 0 END) as nominal_lunas
        FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
        WHERE t.bulan=? GROUP BY n.marketing_nama ORDER BY lunas DESC
    """, [bulan]).fetchall()
    conn.close()
    pct = round(stats["sudah_bayar"]/stats["total_nasabah"]*100,1) if stats["total_nasabah"] else 0
    pesan = f"""📊 *LAPORAN HARIAN BMT AMAL MUSLIM*\n📅 {today}\n\n✅ Sudah Bayar: *{stats['sudah_bayar']} nasabah* ({pct}%)\n⏳ Belum Bayar: *{stats['belum_bayar']} nasabah*\n💰 Terkumpul: *{format_rp(stats['terkumpul'] or 0)}*\n🔴 Tunggakan: *{format_rp(stats['tunggakan'] or 0)}*\n\n💳 Transaksi hari ini: {transaksi['jumlah'] or 0} · {format_rp(transaksi['total'] or 0)}\n\n"""
    for r in marketing_rows:
        pct_m = round(r['lunas']/r['total']*100,0) if r['total'] else 0
        pesan += f"• {r['marketing_nama']}: {r['lunas']}/{r['total']} ({int(pct_m)}%)\n"
    pesan += "\n_Pesan otomatis BMT Billing System_"
    admin_hp = os.environ.get("ADMIN_HP","")
    terkirim = gagal = 0
    if admin_hp:
        result = kirim_wa(admin_hp, pesan)
        if result.get("status") == True: terkirim += 1
        else: gagal += 1
    return jsonify({"success": True, "terkirim": terkirim, "gagal": gagal, "preview_pesan": pesan})

@app.route("/api/notif/reminder-h3", methods=["POST"])
@admin_required
def kirim_reminder_h3():
    import time
    conn = get_db()
    bulan = datetime.now().strftime("%Y-%m")
    today = datetime.now()
    target_hari = today.day  # kirim ke nasabah yang jatuh tempo HARI INI
    rows = conn.execute("""
        SELECT t.id, t.total_tagihan, n.nama, n.no_hp, n.marketing_nama, n.tanggal_jt,
               n.no_rekening, t.angsuran_per_bulan, t.tunggakan_pokok, t.tunggakan_margin
        FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
        WHERE t.bulan=? AND t.status='BELUM' AND n.no_hp IS NOT NULL AND n.no_hp != ''
        AND (n.is_reschedule=0 OR n.is_reschedule IS NULL)
        AND NOT (COALESCE(t.plafond_pokok,0) > 0 AND COALESCE(t.angsuran_per_bulan,0) >= t.plafond_pokok)
    """, [bulan]).fetchall()
    conn.close()
    terkirim = gagal = skip = 0
    delay = int(get_setting('delay_blast_detik', '10'))
    for row in rows:
        tgl = str(row["tanggal_jt"] or "")
        try:
            tgl_num = int(''.join(filter(str.isdigit, tgl[:2])) or 0)
            if tgl_num != target_hari:
                skip += 1; continue
        except:
            skip += 1; continue
        angs = row["angsuran_per_bulan"] or 0
        total = row["total_tagihan"] or 0
        # Formula baru: total > 1 angsuran = ada tunggakan lama
        if angs > 0 and total > angs + max(1000, angs * 0.10):
            actual_tung = round(total - angs)
            nominal = angs
        else:
            actual_tung = 0
            nominal = round(total)
        tgl_fmt = format_tgl_jt(row["tanggal_jt"])
        pesan = pesan_tagihan(row["nama"], nominal, tgl_fmt, row["marketing_nama"],
                              no_akad=row["no_rekening"], tunggakan=actual_tung)
        result = kirim_wa(row["no_hp"], pesan)
        if result.get("status") == True: terkirim += 1
        else: gagal += 1
        if terkirim + gagal < len(rows): time.sleep(max(3, delay))
    return jsonify({"success": True, "terkirim": terkirim, "gagal": gagal, "skip": skip, "target_hari": target_hari})

@app.route("/api/notif/rekap-mingguan", methods=["POST"])
@admin_required
def kirim_rekap_mingguan():
    import time as _time
    conn = get_db()
    bulan = datetime.now().strftime("%Y-%m")
    hari_ini = ["Senin","Selasa","Rabu","Kamis","Jumat","Sabtu","Minggu"][datetime.now().weekday()]
    today_str = datetime.now().strftime("%d/%m/%Y")

    sql_total = (
        "SELECT COUNT(*) as total,"
        " SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN 1 ELSE 0 END) as lunas,"
        " SUM(CASE WHEN (t.status='BELUM') THEN 1 ELSE 0 END) as belum,"
        " SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN t.angsuran_per_bulan ELSE 0 END) as nominal_lunas,"
        " SUM(CASE WHEN (t.status='BELUM' AND t.total_tagihan >= 1) THEN t.total_tagihan ELSE 0 END) as nominal_belum"
        " FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening WHERE t.bulan=?"
    )
    total_row = conn.execute(sql_total, [bulan]).fetchone()

    sql_mkt = (
        "SELECT n.marketing_nama, COUNT(*) as total,"
        " SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN 1 ELSE 0 END) as lunas,"
        " SUM(CASE WHEN (t.status='BELUM') THEN 1 ELSE 0 END) as belum,"
        " SUM(CASE WHEN (t.status IN ('LUNAS','SUDAH_BAYAR')) THEN t.angsuran_per_bulan ELSE 0 END) as nominal_lunas,"
        " SUM(CASE WHEN (t.status='BELUM' AND t.total_tagihan >= 1) THEN t.total_tagihan ELSE 0 END) as nominal_belum"
        " FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening"
        " WHERE t.bulan=? AND n.marketing_nama NOT GLOB '[0-9]*'"
        " GROUP BY n.marketing_nama ORDER BY lunas DESC"
    )
    mkt_rows = conn.execute(sql_mkt, [bulan]).fetchall()

    users = conn.execute(
        "SELECT username, nama, role, no_hp FROM users"
        " WHERE no_hp IS NOT NULL AND no_hp != '' AND no_hp NOT LIKE '%xxxxx%'"
    ).fetchall()
    conn.close()

    def fmt_rp(v):
        v = int(v or 0)
        if v >= 1000000: return "Rp {:.1f}jt".format(v / 1000000)
        if v >= 1000: return "Rp {}rb".format(v // 1000)
        return "Rp {}".format(v)

    pct_total = round(total_row['lunas'] / total_row['total'] * 100, 1) if (total_row and total_row['total']) else 0
    t_lunas  = total_row['lunas']         if total_row else 0
    t_belum  = total_row['belum']         if total_row else 0
    t_total  = total_row['total']         if total_row else 0
    n_lunas  = total_row['nominal_lunas'] if total_row else 0
    n_belum  = total_row['nominal_belum'] if total_row else 0

    pesan_overall = (
        "\U0001f4cb *REKAP MINGGUAN BMT*\n"
        "{}, {}\n\n"
        "*Bulan: {}*\n"
        "Total Nasabah: {}\n"
        "✅ Lunas: {} ({}%)\n"
        "⏳ Belum: {}\n"
        "\U0001f4b0 Terkumpul: {}\n"
        "\U0001f4c9 Sisa Tagihan: {}\n\n"
        "*Per Marketing:*\n"
    ).format(hari_ini, today_str, bulan, t_total, t_lunas, pct_total, t_belum, fmt_rp(n_lunas), fmt_rp(n_belum))

    for r in mkt_rows:
        pct_m = round(r['lunas'] / r['total'] * 100, 0) if r['total'] else 0
        pesan_overall += "• {}: {}/{} ({}%)\n".format(r['marketing_nama'], r['lunas'], r['total'], int(pct_m))
    pesan_overall += "\n_Pesan otomatis BMT Billing System_"

    terkirim = gagal = 0
    delay = max(2, int(get_setting('delay_blast_detik', '10')) // 3)
    mkt_dict = {r['marketing_nama'].upper(): r for r in mkt_rows}
    sent_hps = set()

    admin_hp = os.environ.get("ADMIN_HP", "")
    if admin_hp:
        res = kirim_wa(admin_hp, pesan_overall)
        if res.get("status") is True: terkirim += 1
        else: gagal += 1
        sent_hps.add(admin_hp)

    for usr in users:
        hp = usr['no_hp']
        if hp in sent_hps:
            continue
        sent_hps.add(hp)
        role = usr['role']
        nama = usr['nama']

        if role == 'marketing':
            mkt_data = mkt_dict.get(nama.upper())
            if not mkt_data:
                continue
            pct_m2 = round(mkt_data['lunas'] / mkt_data['total'] * 100, 1) if mkt_data['total'] else 0
            pesan_mkt = (
                "\U0001f4cb *REKAP MINGGUAN - {}*\n"
                "{}, {}\n\n"
                "*Nasabah Anda ({}):*\n"
                "Total: {} nasabah\n"
                "✅ Lunas: {} ({}%)\n"
                "⏳ Belum: {}\n"
                "\U0001f4b0 Terkumpul: {}\n"
                "\U0001f4c9 Sisa: {}\n\n"
                "_Pesan otomatis BMT Billing System_"
            ).format(
                nama, hari_ini, today_str, bulan,
                mkt_data['total'], mkt_data['lunas'], pct_m2, mkt_data['belum'],
                fmt_rp(mkt_data['nominal_lunas']), fmt_rp(mkt_data['nominal_belum'])
            )
            res = kirim_wa(hp, pesan_mkt)
        else:
            # leader / petugas: ringkasan keseluruhan
            res = kirim_wa(hp, pesan_overall)

        if res.get("status") is True: terkirim += 1
        else: gagal += 1
        _time.sleep(delay)

    return jsonify({"success": True, "terkirim": terkirim, "gagal": gagal,
                    "total_marketing": len(mkt_rows), "preview_pesan": pesan_overall})

# ── TOGGLE NOTIF ───────────────────────────────────────────────
@app.route("/api/notif/status", methods=["GET"])
@admin_required
def get_notif_status():
    flag_path = os.path.join("data", "notif_flag.txt")
    try:
        with open(flag_path, "r") as ff:
            flag = ff.read().strip()
    except:
        flag = os.environ.get("NOTIF_AKTIF", "0")
    return jsonify({"aktif": flag == "1"})

@app.route("/api/notif/toggle", methods=["POST"])
@admin_required
def toggle_notif():
    data = request.json
    aktif = "1" if data.get("aktif") else "0"
    flag_path = os.path.join("data", "notif_flag.txt")
    with open(flag_path, "w") as ff:
        ff.write(aktif)
    return jsonify({"success": True, "aktif": aktif == "1"})


# FOTO KUNJUNGAN
@app.route("/foto_kunjungan/<path:filename>")
@login_required
def serve_foto_kunjungan(filename):
    filename = os.path.basename(filename)
    from flask import send_from_directory
    import os as _os
    return send_from_directory(_os.path.join(_os.getcwd(), "data/foto_kunjungan"), filename)


# ── BLAST LOG ──────────────────────────────────────────────────────────────
def ensure_blast_log(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS blast_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipe TEXT NOT NULL,
            bulan TEXT NOT NULL,
            dilakukan_oleh TEXT,
            terkirim INTEGER DEFAULT 0,
            gagal INTEGER DEFAULT 0,
            skip INTEGER DEFAULT 0,
            catatan TEXT,
            dibuat_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS blast_tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id TEXT UNIQUE NOT NULL,
            tipe TEXT NOT NULL,
            bulan TEXT NOT NULL,
            status TEXT DEFAULT 'running',
            terkirim INTEGER DEFAULT 0,
            gagal INTEGER DEFAULT 0,
            total INTEGER DEFAULT 0,
            dilakukan_oleh TEXT,
            catatan TEXT,
            dibuat_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            selesai_at TIMESTAMP
        )
    """)
    conn.commit()

def ensure_wa_send_log(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS wa_send_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            no_rekening TEXT NOT NULL,
            nama TEXT,
            no_hp TEXT,
            bulan TEXT NOT NULL,
            tipe TEXT NOT NULL,
            status TEXT DEFAULT 'terkirim',
            dikirim_oleh TEXT,
            dikirim_at TIMESTAMP DEFAULT (datetime('now','localtime'))
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_wa_send_bulan ON wa_send_log(bulan, no_rekening)")
    conn.commit()

@app.route("/api/blast/log")
@admin_required
def get_blast_log():
    conn = get_db()
    ensure_blast_log(conn)
    bulan = request.args.get("bulan", "")
    limit = int(request.args.get("limit", "20"))
    if bulan:
        rows = conn.execute(
            "SELECT * FROM blast_log WHERE bulan=? ORDER BY dibuat_at DESC LIMIT ?",
            [bulan, limit]
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM blast_log ORDER BY dibuat_at DESC LIMIT ?",
            [limit]
        ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/blast/log/check")
@admin_required
def check_blast_log():
    """Cek apakah bulan tertentu sudah pernah di-blast"""
    conn = get_db()
    ensure_blast_log(conn)
    bulan = request.args.get("bulan", datetime.now().strftime("%Y-%m"))
    rows = conn.execute(
        "SELECT * FROM blast_log WHERE bulan=? AND tipe='execute_blast' ORDER BY dibuat_at DESC LIMIT 5",
        [bulan]
    ).fetchall()
    conn.close()
    return jsonify({"sudah_blast": len(rows) > 0, "history": [dict(r) for r in rows]})

@app.route("/api/wa-log")
@admin_required
def get_wa_log():
    bulan = request.args.get("bulan", datetime.now().strftime("%Y-%m"))
    conn = get_db()
    ensure_wa_send_log(conn)
    rows = conn.execute(
        "SELECT * FROM wa_send_log WHERE bulan=? ORDER BY dikirim_at DESC LIMIT 200",
        [bulan]
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/wa-log/bulan-ini")
@login_required
def wa_sent_this_month():
    bulan = request.args.get("bulan", datetime.now().strftime("%Y-%m"))
    conn = get_db()
    ensure_wa_send_log(conn)
    rows = conn.execute(
        "SELECT no_rekening, COUNT(*) as jumlah FROM wa_send_log WHERE bulan=? AND status='terkirim' GROUP BY no_rekening",
        [bulan]
    ).fetchall()
    conn.close()
    return jsonify({r["no_rekening"]: r["jumlah"] for r in rows})

def ensure_kunjungan_table(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS kunjungan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            no_rekening TEXT,
            bulan TEXT,
            tanggal TEXT,
            catatan TEXT,
            foto_path TEXT,
            marketing_id TEXT,
            dicatat_oleh TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    # Kolom pembayaran kunjungan (utk ukur prestasi petugas, TIDAK mengubah tagihan/MESPro)
    _cols = [r[1] for r in conn.execute("PRAGMA table_info(kunjungan)").fetchall()]
    if "nominal_bayar" not in _cols:
        conn.execute("ALTER TABLE kunjungan ADD COLUMN nominal_bayar REAL DEFAULT 0")
    if "cara_bayar" not in _cols:
        conn.execute("ALTER TABLE kunjungan ADD COLUMN cara_bayar TEXT")
    if "janji_tanggal" not in _cols:
        conn.execute("ALTER TABLE kunjungan ADD COLUMN janji_tanggal TEXT")
    if "foto_hash" not in _cols:
        conn.execute("ALTER TABLE kunjungan ADD COLUMN foto_hash TEXT")
    conn.commit()

# MONITORING KOLEKTIBILITAS 2-5
@app.route("/api/monitoring/nasabah", methods=["GET"])
@login_required
def monitoring_nasabah():
    bulan = request.args.get("bulan", datetime.now().strftime("%Y-%m"))
    conn = get_db()
    ensure_kunjungan_table(conn)
    marketing_id = session.get("marketing_id")
    role = get_user_role()
    params = [bulan]
    extra_where = ""
    if role not in ("admin","leader","petugas") and marketing_id:
        extra_where = " AND n.marketing_id=?"
        params.append(marketing_id)
    reschedule = request.args.get("reschedule", "")
    if reschedule == "1":
        kol_filter = "AND n.is_reschedule=1"
    else:
        kol_filter = "AND (t.kolektibilitas >= 2 OR n.is_reschedule=1)"
    rows = conn.execute(f"""
        SELECT t.id as tagihan_id, t.no_rekening, t.total_tagihan, t.kolektibilitas,
               t.tunggakan_pokok, t.tunggakan_margin, t.status,
               n.nama, n.no_hp, n.marketing_nama, n.tanggal_jt,
               n.is_reschedule, n.tgl_realisasi
        FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
        WHERE t.bulan=? {kol_filter}{extra_where}
        ORDER BY t.kolektibilitas DESC, n.marketing_nama, n.nama
    """, params).fetchall()
    kj = conn.execute(
        "SELECT no_rekening, COUNT(*) as jumlah FROM kunjungan WHERE bulan=? GROUP BY no_rekening",
        [bulan]).fetchall()
    kj_map = {r["no_rekening"]: r["jumlah"] for r in kj}
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        d["jumlah_kunjungan"] = kj_map.get(r["no_rekening"], 0)
        result.append(d)
    return jsonify(result)

@app.route("/api/kunjungan", methods=["POST"])
@login_required
def tambah_kunjungan():
    import uuid
    no_rekening = request.form.get("no_rekening", "")
    bulan = request.form.get("bulan", datetime.now().strftime("%Y-%m"))
    catatan = request.form.get("catatan", "").strip()
    foto_path = None
    foto_hash = None
    if "foto" in request.files:
        foto = request.files["foto"]
        if foto and foto.filename:
            ext = foto.filename.rsplit(".", 1)[-1].lower() if "." in foto.filename else "jpg"
            if ext not in ("jpg", "jpeg", "png", "webp", "heic", "gif", "bmp"):
                return jsonify({"error": "Format foto tidak didukung"}), 400
            filename = "{}_{}_{}.jpg".format(
                no_rekening, datetime.now().strftime("%Y%m%d%H%M%S"), uuid.uuid4().hex[:6])
            compressed = compress_image(foto.stream)
            # Hash sidik-jari foto (utk deteksi foto double di audit kunjungan)
            try:
                from audit_kunjungan import hitung_hash_dari_buf
                foto_hash = hitung_hash_dari_buf(compressed)
            except Exception:
                foto_hash = None
            file_id = upload_to_gdrive(compressed, filename)
            if file_id:
                foto_path = "gdrive:" + file_id
            else:
                foto.stream.seek(0)
                foto.save(os.path.join("data/foto_kunjungan", filename))
                foto_path = filename
    try:
        nominal_bayar = float(request.form.get("nominal_bayar", 0) or 0)
    except (ValueError, TypeError):
        nominal_bayar = 0
    cara_bayar = (request.form.get("cara_bayar", "") or "").strip() or None
    janji_tanggal = (request.form.get("janji_tanggal", "") or "").strip() or None
    conn = get_db()
    ensure_kunjungan_table(conn)
    conn.execute("""
        INSERT INTO kunjungan (no_rekening, bulan, tanggal, catatan, foto_path,
                               marketing_id, dicatat_oleh, nominal_bayar, cara_bayar, janji_tanggal,
                               foto_hash)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
    """, (no_rekening, bulan, datetime.now().strftime("%Y-%m-%d"), catatan,
          foto_path, session.get("marketing_id"), session.get("nama"),
          nominal_bayar, cara_bayar, janji_tanggal, foto_hash))
    conn.commit()
    conn.close()
    # CATATAN: pembayaran kunjungan TIDAK mengubah status/angka tagihan.
    # Patokan tagihan tetap dari MESPro. Ini hanya catatan prestasi petugas.
    return jsonify({"success": True, "foto_path": foto_path, "nominal_bayar": nominal_bayar})

@app.route("/api/kunjungan/audit", methods=["GET"])
@login_required
def audit_kunjungan_route():
    """Audit kunjungan (rule-based) — hanya admin & leader.
    Menandai kunjungan yang PERLU DIPERIKSA (foto double, catatan duplikat,
    input menumpuk, backdating, dll). Output = indikasi, bukan vonis."""
    role = get_user_role()
    if role not in ("admin", "leader"):
        return jsonify({"error": "Akses ditolak — hanya admin/leader"}), 403
    bulan = request.args.get("bulan", datetime.now().strftime("%Y-%m"))
    conn = get_db()
    ensure_kunjungan_table(conn)
    try:
        from audit_kunjungan import audit_bulan
        hasil = audit_bulan(conn, bulan)
    except Exception as e:
        conn.close()
        return jsonify({"error": "Audit gagal: " + str(e)}), 500
    conn.close()
    return jsonify(hasil)

@app.route("/api/kunjungan/<no_rekening>", methods=["GET"])
@login_required
def get_kunjungan(no_rekening):
    bulan = request.args.get("bulan", datetime.now().strftime("%Y-%m"))
    conn = get_db()
    ensure_kunjungan_table(conn)
    rows = conn.execute(
        "SELECT * FROM kunjungan WHERE no_rekening=? AND bulan=? ORDER BY created_at DESC",
        [no_rekening, bulan]).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])



@app.route("/api/janji-bayar")
@login_required
def janji_bayar():
    """Janji bayar nasabah (dari kunjungan) yg belum dipenuhi -> pengingat petugas/leader/admin."""
    conn = get_db(); ensure_kunjungan_table(conn)
    role = get_user_role(); marketing_id = session.get("marketing_id")
    bulan = datetime.now().strftime("%Y-%m")
    today = datetime.now().strftime("%Y-%m-%d")
    mkt = ""; params = [bulan]
    if role not in ("admin", "leader", "petugas") and marketing_id:
        mkt = " AND n.marketing_id=?"; params.append(marketing_id)
    rows = conn.execute(f"""
        SELECT k.no_rekening, n.nama, n.no_hp, n.marketing_nama, k.janji_tanggal,
               k.tanggal AS tgl_kunjungan, k.catatan, k.dicatat_oleh,
               t.status, t.total_tagihan, t.kolektibilitas
        FROM kunjungan k
        JOIN nasabah n ON k.no_rekening = n.no_rekening
        LEFT JOIN tagihan t ON t.no_rekening = k.no_rekening AND t.bulan = ?
        WHERE k.janji_tanggal IS NOT NULL AND k.janji_tanggal != ''
          AND k.id IN (SELECT MAX(id) FROM kunjungan
                       WHERE janji_tanggal IS NOT NULL AND janji_tanggal != ''
                       GROUP BY no_rekening)
          {mkt}
        ORDER BY k.janji_tanggal ASC
    """, params).fetchall()
    data = []
    for r in rows:
        d = dict(r)
        # janji yg sudah dipenuhi (sudah bayar) tidak perlu diingatkan
        if d.get("status") in ("LUNAS", "SUDAH_BAYAR"):
            continue
        data.append(d)
    conn.close()
    return jsonify({"data": data, "today": today, "total": len(data)})

@app.route("/api/npl/target")
@login_required
def npl_target():
    """Target penurunan NPL (kol 2-5) menuju <5% akhir tahun + watchlist per-nasabah."""
    bulan = request.args.get("bulan", datetime.now().strftime("%Y-%m"))
    y, m = map(int, bulan.split("-"))
    pm = f"{y-1}-12" if m == 1 else f"{y}-{m-1:02d}"
    conn = get_db()
    role = get_user_role(); marketing_id = session.get("marketing_id")
    mkt = ""; base = [bulan]
    if role not in ("admin", "leader", "petugas") and marketing_id:
        mkt = " AND n.marketing_id=?"; base.append(marketing_id)
    tot = conn.execute(f"""
        SELECT COALESCE(SUM(t.saldo_pinjaman),0) s,
               COALESCE(SUM(CASE WHEN t.kolektibilitas>=2 THEN t.saldo_pinjaman ELSE 0 END),0) npl,
               COUNT(CASE WHEN t.kolektibilitas>=2 THEN 1 END) npl_cnt
        FROM tagihan t JOIN nasabah n ON t.no_rekening=n.no_rekening
        WHERE t.bulan=?{mkt}""", base).fetchone()
    total_saldo = tot["s"] or 0; npl_saldo = tot["npl"] or 0
    npl_pct = round(npl_saldo/total_saldo*100, 2) if total_saldo else 0
    target_pct = 5.0
    max_allowed = total_saldo * target_pct / 100.0
    excess = max(0.0, npl_saldo - max_allowed)
    sisa_bulan = max(1, 12 - m + 1)
    per_bulan = excess / sisa_bulan
    traj = []
    span = 12 - m
    for mm in range(m, 13):
        frac = (mm - m) / span if span > 0 else 1
        traj.append({"bulan": f"{y}-{mm:02d}",
                     "target_pct": round(npl_pct + (target_pct - npl_pct) * frac, 2)})
    rows = conn.execute(f"""
        SELECT n.nama, n.no_rekening, n.marketing_nama, n.no_hp,
               t.kolektibilitas, t.saldo_pinjaman, t.total_tagihan, t.angsuran_per_bulan, n.is_reschedule,
               (SELECT kolektibilitas FROM tagihan tp WHERE tp.no_rekening=t.no_rekening AND tp.bulan=?) as kol_prev
        FROM tagihan t JOIN nasabah n ON t.no_rekening=n.no_rekening
        WHERE t.bulan=? AND t.kolektibilitas>=2{mkt}
        ORDER BY t.kolektibilitas DESC, t.saldo_pinjaman DESC""", [pm]+base).fetchall()
    watch = []
    membaik = memburuk = tetap = 0
    for r in rows:
        d = dict(r)
        kp = d.get("kol_prev")
        if kp is None: d["trend"] = "baru"
        elif d["kolektibilitas"] < kp: d["trend"] = "membaik"; membaik += 1
        elif d["kolektibilitas"] > kp: d["trend"] = "memburuk"; memburuk += 1
        else: d["trend"] = "tetap"; tetap += 1
        d["target_kol"] = max(1, d["kolektibilitas"] - 1)
        _allow = {1: 0, 2: 2, 3: 3, 4: 6}
        ang = d.get("angsuran_per_bulan") or 0
        tung = d.get("total_tagihan") or 0
        bayar = tung - _allow.get(d["target_kol"], 0) * ang
        if bayar < 0:
            bayar = 0
        if bayar <= 0 and tung > 0:
            bayar = ang if ang > 0 else tung
        d["target_bayar"] = round(bayar)
        watch.append(d)
    conn.close()
    return jsonify({
        "bulan": bulan, "total_saldo": total_saldo, "npl_saldo": npl_saldo,
        "npl_pct": npl_pct, "target_pct": target_pct, "max_allowed": max_allowed,
        "excess": excess, "sisa_bulan": sisa_bulan, "per_bulan": per_bulan,
        "trajectory": traj, "watchlist": watch, "jml_watchlist": len(watch),
        "trend": {"membaik": membaik, "memburuk": memburuk, "tetap": tetap},
        "tercapai": npl_pct < target_pct
    })

@app.route("/api/kunjungan/rekap")
@login_required
def kunjungan_rekap():
    """Rekap kunjungan untuk ukur prestasi petugas (harian/bulanan).
    Ringkasan per petugas + detail tiap kunjungan."""
    bulan = request.args.get("bulan")
    tanggal = request.args.get("tanggal")
    conn = get_db()
    ensure_kunjungan_table(conn)
    role = get_user_role(); marketing_id = session.get("marketing_id")
    where = []; params = []
    if tanggal:
        where.append("k.tanggal = ?"); params.append(tanggal)
    else:
        if not bulan:
            bulan = datetime.now().strftime("%Y-%m")
        where.append("k.bulan = ?"); params.append(bulan)
    if role not in ("admin", "leader", "petugas") and marketing_id:
        where.append("k.marketing_id = ?"); params.append(marketing_id)
    wsql = "WHERE " + " AND ".join(where)
    rows = conn.execute(f"""
        SELECT k.id, k.no_rekening, k.tanggal, k.catatan, k.foto_path,
               k.dicatat_oleh, k.nominal_bayar, k.cara_bayar,
               n.nama, n.alamat, n.marketing_nama, t.kolektibilitas
        FROM kunjungan k
        LEFT JOIN nasabah n ON k.no_rekening = n.no_rekening
        LEFT JOIN tagihan t ON t.no_rekening = k.no_rekening AND t.bulan = k.bulan
        {wsql}
        ORDER BY k.tanggal DESC, k.id DESC
    """, params).fetchall()
    detail = [dict(r) for r in rows]
    summ = {}
    for d in detail:
        p = d["dicatat_oleh"] or "-"
        s = summ.setdefault(p, {"petugas": p, "jml_kunjungan": 0,
                                "_nasabah": set(), "total_nominal": 0.0})
        s["jml_kunjungan"] += 1
        s["_nasabah"].add(d["no_rekening"])
        s["total_nominal"] += (d["nominal_bayar"] or 0)
    summary = [{"petugas": s["petugas"], "jml_kunjungan": s["jml_kunjungan"],
                "jml_nasabah": len(s["_nasabah"]), "total_nominal": s["total_nominal"]}
               for s in summ.values()]
    summary.sort(key=lambda x: (-x["total_nominal"], -x["jml_kunjungan"]))
    conn.close()
    return jsonify({
        "summary": summary, "detail": detail,
        "total_kunjungan": len(detail),
        "total_nasabah": len(set(d["no_rekening"] for d in detail)),
        "total_nominal": sum(d["nominal_bayar"] or 0 for d in detail),
        "jml_petugas": len(summary),
        "bulan": bulan, "tanggal": tanggal
    })

@app.route("/api/kunjungan/rekap/export-xlsx")
@login_required
def kunjungan_rekap_export_xlsx():
    """Export Excel (.xlsx) rekap kunjungan bulanan: 1 baris per nasabah,
    kolom kunjungan 1-5 + link foto. Hanya admin/leader."""
    role = get_user_role()
    if role not in ("admin", "leader"):
        return jsonify({"error": "Akses ditolak"}), 403
    bulan = request.args.get("bulan") or datetime.now().strftime("%Y-%m")
    conn = get_db()
    ensure_kunjungan_table(conn)
    rows = conn.execute("""
        SELECT k.id, k.no_rekening, k.tanggal, k.catatan, k.foto_path,
               k.dicatat_oleh, k.nominal_bayar, k.cara_bayar,
               n.nama, n.alamat, n.marketing_nama, t.kolektibilitas
        FROM kunjungan k
        LEFT JOIN nasabah n ON k.no_rekening = n.no_rekening
        LEFT JOIN tagihan t ON t.no_rekening = k.no_rekening AND t.bulan = k.bulan
        WHERE k.bulan = ?
        ORDER BY k.tanggal ASC, k.id ASC
    """, [bulan]).fetchall()
    conn.close()

    from collections import OrderedDict
    groups = OrderedDict()
    for r in rows:
        key = r["no_rekening"] or "-"
        g = groups.get(key)
        if not g:
            g = {"no_rekening": key, "nama": r["nama"], "alamat": r["alamat"],
                 "kol": r["kolektibilitas"], "petugas": [], "visits": [], "total_bayar": 0.0}
            groups[key] = g
        g["visits"].append(r)
        g["total_bayar"] += (r["nominal_bayar"] or 0)
        if r["dicatat_oleh"] and r["dicatat_oleh"] not in g["petugas"]:
            g["petugas"].append(r["dicatat_oleh"])
        if not g["kol"] and r["kolektibilitas"]:
            g["kol"] = r["kolektibilitas"]

    host = request.host_url.rstrip("/")
    def foto_url(fp):
        if not fp:
            return None
        if fp.startswith("gdrive:"):
            return "https://drive.google.com/thumbnail?id=" + fp.replace("gdrive:", "") + "&sz=w800"
        return host + "/foto_kunjungan/" + fp

    def rupiah(n):
        try:
            return "Rp" + "{:,.0f}".format(n).replace(",", ".")
        except Exception:
            return "Rp0"

    def visit_text(v):
        parts = []
        if v["tanggal"]:
            parts.append(str(v["tanggal"]))
        if v["catatan"]:
            parts.append(v["catatan"])
        if (v["nominal_bayar"] or 0) > 0:
            parts.append("Bayar " + rupiah(v["nominal_bayar"]) +
                         ((" " + v["cara_bayar"]) if v["cara_bayar"] else ""))
        return "\n".join(parts)

    import openpyxl, io as _io
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    KL = ["", "Lancar", "DPK", "Kurang Lancar", "Diragukan", "Macet"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Kunjungan"
    headers = ["No", "Nama", "Alamat", "No. Akad", "Kol", "Petugas", "Jml Kunjungan",
               "Kunjungan 1", "Kunjungan 2", "Kunjungan 3", "Kunjungan 4", "Kunjungan 5",
               "Jumlah Bayar (Rp)", "Foto 1", "Foto 2", "Foto 3", "Foto 4", "Foto 5"]
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="047857")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    for c in ws[1]:
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
    link_font = Font(color="0563C1", underline="single")

    rownum = 2
    for idx, g in enumerate(groups.values(), start=1):
        visits = g["visits"]
        row = [idx, g["nama"] or "-", g["alamat"] or "-", g["no_rekening"] or "-",
               (KL[g["kol"]] if g["kol"] else "-"),
               ", ".join(g["petugas"]) or "-", len(visits)]
        for i in range(5):
            row.append(visit_text(visits[i]) if i < len(visits) else "")
        row.append(round(g["total_bayar"]))
        for i in range(5):
            row.append("")  # foto placeholder, diisi hyperlink di bawah
        ws.append(row)
        for i in range(5):
            if i < len(visits) and visits[i]["foto_path"]:
                cell = ws.cell(row=rownum, column=14 + i)
                cell.value = "Foto %d" % (i + 1)
                cell.hyperlink = foto_url(visits[i]["foto_path"])
                cell.font = link_font
        for col in range(8, 13):
            ws.cell(row=rownum, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        rownum += 1

    widths = [4, 22, 28, 16, 14, 14, 11, 26, 26, 26, 26, 26, 15, 8, 8, 8, 8, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Baris TOTAL di bawah data utama
    total_kunj = sum(len(g["visits"]) for g in groups.values())
    total_bayar_all = sum(g["total_bayar"] for g in groups.values())
    total_fill = PatternFill("solid", fgColor="D1FAE5")
    ws.append(["", "TOTAL (%d nasabah)" % len(groups), "", "", "", "", total_kunj,
               "", "", "", "", "", round(total_bayar_all), "", "", "", "", ""])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
        c.fill = total_fill

    # Sheet kedua: ringkasan per petugas
    pet = OrderedDict()
    for r in rows:
        p = r["dicatat_oleh"] or "-"
        d = pet.get(p)
        if not d:
            d = {"kunj": 0, "nasabah": set(), "nominal": 0.0}
            pet[p] = d
        d["kunj"] += 1
        d["nasabah"].add(r["no_rekening"])
        d["nominal"] += (r["nominal_bayar"] or 0)
    ws2 = wb.create_sheet("Ringkasan Petugas")
    ws2.append(["Petugas", "Jml Kunjungan", "Jml Nasabah", "Total Setoran (Rp)"])
    for c in ws2[1]:
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    for p, d in sorted(pet.items(), key=lambda kv: (-kv[1]["nominal"], -kv[1]["kunj"])):
        ws2.append([p, d["kunj"], len(d["nasabah"]), round(d["nominal"])])
    ws2.append(["TOTAL", total_kunj, len(groups), round(total_bayar_all)])
    for c in ws2[ws2.max_row]:
        c.font = Font(bold=True)
        c.fill = total_fill
    for i, w in enumerate([24, 14, 13, 18], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import Response
    fname = "rekap_kunjungan_%s.xlsx" % bulan
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=" + fname})

@app.route("/api/kunjungan/<int:kj_id>", methods=["DELETE"])
@login_required
def hapus_kunjungan(kj_id):
    if get_user_role() != "admin":
        return jsonify({"error": "Akses ditolak"}), 403
    conn = get_db()
    row = conn.execute("SELECT foto_path FROM kunjungan WHERE id=?", [kj_id]).fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "Data tidak ditemukan"}), 404
    conn.execute("DELETE FROM kunjungan WHERE id=?", [kj_id])
    conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route("/api/kunjungan/rekap-harian", methods=["GET"])
@login_required
def rekap_harian_kunjungan():
    tanggal = request.args.get("tanggal", datetime.now().strftime("%Y-%m-%d"))
    bulan = tanggal[:7]
    conn = get_db()
    marketing_id_k = session.get("marketing_id")
    role_k = get_user_role()
    params_k = [bulan, tanggal]
    extra_k = ""
    if role_k not in ("admin","leader","petugas") and marketing_id_k:
        extra_k = " AND k.marketing_id=?"
        params_k.append(marketing_id_k)
    ensure_kunjungan_table(conn)
    rows = conn.execute(f"""
        SELECT k.id, k.no_rekening, k.tanggal, k.catatan, k.foto_path,
               k.dicatat_oleh, k.marketing_id,
               n.nama, n.marketing_nama,
               t.kolektibilitas, t.total_tagihan,
               t.tunggakan_pokok, t.tunggakan_margin,
               CASE
                   WHEN t.status = 'LUNAS' THEN 'LUNAS'
                   WHEN EXISTS (
                       SELECT 1 FROM pembayaran p2
                       WHERE p2.no_rekening = k.no_rekening
                         AND DATE(p2.tanggal) = k.tanggal
                   ) THEN 'LUNAS'
                   ELSE COALESCE(t.status, 'BELUM')
               END as status
        FROM kunjungan k
        JOIN nasabah n ON k.no_rekening = n.no_rekening
        LEFT JOIN tagihan t ON t.no_rekening = k.no_rekening AND t.bulan = ?
        WHERE k.tanggal = ?{extra_k}
        ORDER BY k.dicatat_oleh, n.nama
    """, params_k).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/monitoring/rekap", methods=["GET"])
@login_required
def monitoring_rekap():
    bulan = request.args.get("bulan", datetime.now().strftime("%Y-%m"))
    conn = get_db()
    ensure_kunjungan_table(conn)
    marketing_id = session.get("marketing_id")
    role = get_user_role()
    params_r = [bulan]
    extra_r = ""
    if role not in ("admin","leader","petugas") and marketing_id:
        extra_r = " AND n.marketing_id=?"
        params_r.append(marketing_id)
    rows = conn.execute(f"""
        SELECT t.id as tagihan_id, t.no_rekening, n.nama, n.marketing_nama, t.kolektibilitas,
               t.total_tagihan, t.tunggakan_pokok, t.tunggakan_margin, t.status, n.tanggal_jt,
               n.is_reschedule
        FROM tagihan t JOIN nasabah n ON t.no_rekening = n.no_rekening
        WHERE t.bulan=? AND (t.kolektibilitas >= 2 OR n.is_reschedule=1){extra_r}
        ORDER BY t.kolektibilitas DESC, n.marketing_nama, n.nama
    """, params_r).fetchall()
    kj = conn.execute("""
        SELECT no_rekening, COUNT(*) as jumlah, MAX(tanggal) as terakhir,
               GROUP_CONCAT(catatan, ' || ') as catatan_all
        FROM kunjungan WHERE bulan=? GROUP BY no_rekening
    """, [bulan]).fetchall()
    kj_map = {r["no_rekening"]: dict(r) for r in kj}
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        k = kj_map.get(r["no_rekening"], {})
        d["jumlah_kunjungan"] = k.get("jumlah", 0)
        d["terakhir_kunjungan"] = k.get("terakhir") or "-"
        d["catatan_kunjungan"] = k.get("catatan_all") or ""
        result.append(d)
    return jsonify(result)

# PWA routes
@app.route('/manifest.json')
def pwa_manifest():
    return send_from_directory('static', 'manifest.json', mimetype='application/manifest+json')

@app.route('/sw.js')
def service_worker():
    resp = make_response(send_from_directory('static', 'sw.js'))
    resp.headers['Service-Worker-Allowed'] = '/'
    resp.headers['Cache-Control'] = 'no-cache'
    return resp
