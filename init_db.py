import sqlite3
import os
import openpyxl
from datetime import datetime

DB_PATH = "data/koperasi.db"

def init_db():
    os.makedirs("data", exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # Tabel users
    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            nama TEXT NOT NULL,
            role TEXT DEFAULT 'marketing',
            marketing_id TEXT,
            aktif INTEGER DEFAULT 1
        )
    """)

    # Tabel nasabah
    c.execute("""
        CREATE TABLE IF NOT EXISTS nasabah (
            no_rekening TEXT PRIMARY KEY,
            nama TEXT,
            no_hp TEXT,
            marketing_id TEXT,
            marketing_nama TEXT,
            tanggal_jt TEXT,
            aktif INTEGER DEFAULT 1
        )
    """)

    # Tabel tagihan
    c.execute("""
        CREATE TABLE IF NOT EXISTS tagihan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            no_rekening TEXT,
            bulan TEXT,
            saldo_pinjaman REAL DEFAULT 0,
            tunggakan_pokok REAL DEFAULT 0,
            tunggakan_margin REAL DEFAULT 0,
            total_tagihan REAL DEFAULT 0,
            kolektibilitas INTEGER DEFAULT 1,
            status TEXT DEFAULT 'BELUM',
            cara_bayar TEXT,
            tgl_angsuran TEXT,
            keterangan TEXT,
            UNIQUE(no_rekening, bulan)
        )
    """)

    # Tabel pembayaran
    c.execute("""
        CREATE TABLE IF NOT EXISTS pembayaran (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            no_rekening TEXT,
            tagihan_id INTEGER,
            jumlah REAL,
            cara_bayar TEXT,
            marketing_id TEXT,
            catatan TEXT,
            dicatat_oleh TEXT,
            tanggal TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Tabel import_log
    c.execute("""
        CREATE TABLE IF NOT EXISTS import_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bulan TEXT,
            filepath TEXT,
            nasabah_baru INTEGER DEFAULT 0,
            nasabah_update INTEGER DEFAULT 0,
            nasabah_nonaktif INTEGER DEFAULT 0,
            tagihan_baru INTEGER DEFAULT 0,
            tagihan_update INTEGER DEFAULT 0,
            diimport_oleh TEXT,
            waktu TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Default admin user (password: admin123)
    import hashlib
    default_pw = hashlib.sha256("admin123".encode()).hexdigest()
    c.execute("""
        INSERT OR IGNORE INTO users (username, password, nama, role, marketing_id)
        VALUES ('admin', ?, 'Administrator', 'admin', NULL)
    """, (default_pw,))

    # Default marketing users (password: bmt2026)
    default_pw_mkt = hashlib.sha256("bmt2026".encode()).hexdigest()
    marketing_list = [
        ("raafi",    "RAAFI",    "MKT001"),
        ("rizal",    "RIZAL",    "MKT002"),
        ("nikko",    "NIKKO",    "MKT003"),
        ("lilik",    "LILIK",    "MKT004"),
        ("widi",     "WIDI",     "MKT005"),
        ("siswanto", "SISWANTO", "MKT006"),
        ("suratman", "SURATMAN", "MKT007"),
        ("tumino",   "TUMINO",   "MKT008"),
        ("agus",     "AGUS",     "MKT009"),
        ("sdit",     "SDIT",     "MKT010"),
        ("edhi",     "EDHI",     "MKT011"),
        ("joko",     "JOKO",     "MKT012"),
        ("ekomey",   "EKO MEY",  "MKT013"),
        ("bam",      "BAM",      "MKT014"),
        ("lilis",    "LILIS",    "MKT015"),
        ("aguss",    "AGUS S",   "MKT016"),
    ]
    for username, nama, mkt_id in marketing_list:
        c.execute("""
            INSERT OR IGNORE INTO users (username, password, nama, role, marketing_id)
            VALUES (?, ?, ?, 'marketing', ?)
        """, (username, default_pw_mkt, nama, mkt_id))

    conn.commit()
    conn.close()
    print("✅ Database berhasil diinisialisasi!")


def import_excel(filepath, diimport_oleh="admin"):
    """
    Import data nasabah & tagihan dari file Excel.
    Kolom yang diharapkan (nama kolom fleksibel, pakai pencocokan keyword):
    no_rekening, nama, no_hp, marketing_id, marketing_nama, tanggal_jt,
    saldo_pinjaman, tunggakan_pokok, tunggakan_margin, total_tagihan, kolektibilitas
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        # Baca header baris pertama
        headers = []
        for cell in ws[1]:
            val = str(cell.value).strip().lower() if cell.value else ""
            headers.append(val)

        def find_col(keywords):
            for kw in keywords:
                for i, h in enumerate(headers):
                    if kw in h:
                        return i
            return None

        # Mapping kolom
        col = {
            "no_rek":       find_col(["no_rek", "rekening", "no rekening"]),
            "nama":         find_col(["nama"]),
            "no_hp":        find_col(["hp", "telp", "telepon", "handphone"]),
            "mkt_id":       find_col(["marketing_id", "kode marketing", "id marketing"]),
            "mkt_nama":     find_col(["marketing_nama", "nama marketing", "marketing"]),
            "tgl_jt":       find_col(["jatuh tempo", "tanggal_jt", "tgl_jt", "jt"]),
            "saldo":        find_col(["saldo_pinjaman", "saldo pinjaman", "saldo"]),
            "tung_pokok":   find_col(["tunggakan_pokok", "tunggakan pokok", "pokok"]),
            "tung_margin":  find_col(["tunggakan_margin", "tunggakan margin", "margin"]),
            "total":        find_col(["total_tagihan", "total tagihan", "total"]),
            "kolek":        find_col(["kolektibilitas", "kolek", "kol"]),
        }

        # Deteksi bulan dari nama file (format: ...MEI 2026... atau ...2026-05...)
        import re
        bulan_map = {
            "januari":"01","februari":"02","maret":"03","april":"04",
            "mei":"05","juni":"06","juli":"07","agustus":"08",
            "september":"09","oktober":"10","november":"11","desember":"12"
        }
        bulan = datetime.now().strftime("%Y-%m")
        fname = os.path.basename(filepath).lower()
        for nama_bln, num in bulan_map.items():
            if nama_bln in fname:
                tahun_match = re.search(r'20\d{2}', fname)
                tahun = tahun_match.group() if tahun_match else str(datetime.now().year)
                bulan = f"{tahun}-{num}"
                break

        conn = sqlite3.connect(DB_PATH)
        nasabah_baru = nasabah_update = nasabah_nonaktif = 0
        tagihan_baru = tagihan_update = 0

        no_rek_excel = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            def val(c):
                if c is None:
                    return None
                v = row[c]
                return str(v).strip() if v is not None else None

            def num(c):
                if c is None:
                    return 0
                v = row[c]
                try:
                    return float(str(v).replace(",", "").replace(".", "")) if v else 0
                except:
                    return 0

            no_rek = val(col["no_rek"])
            if not no_rek or no_rek.lower() in ("none", ""):
                continue

            no_rek_excel.add(no_rek)

            # UPSERT nasabah
            existing = conn.execute(
                "SELECT no_rekening FROM nasabah WHERE no_rekening=?", (no_rek,)
            ).fetchone()

            if existing:
                conn.execute("""
                    UPDATE nasabah SET nama=?, no_hp=?, marketing_id=?, marketing_nama=?,
                    tanggal_jt=?, aktif=1 WHERE no_rekening=?
                """, (val(col["nama"]), val(col["no_hp"]), val(col["mkt_id"]),
                      val(col["mkt_nama"]), val(col["tgl_jt"]), no_rek))
                nasabah_update += 1
            else:
                conn.execute("""
                    INSERT INTO nasabah (no_rekening, nama, no_hp, marketing_id, marketing_nama, tanggal_jt)
                    VALUES (?,?,?,?,?,?)
                """, (no_rek, val(col["nama"]), val(col["no_hp"]),
                      val(col["mkt_id"]), val(col["mkt_nama"]), val(col["tgl_jt"])))
                nasabah_baru += 1

            # UPSERT tagihan
            saldo       = num(col["saldo"])
            tung_pokok  = num(col["tung_pokok"])
            tung_margin = num(col["tung_margin"])
            total       = num(col["total"]) or (tung_pokok + tung_margin)
            kolek       = int(num(col["kolek"])) if col["kolek"] else 1

            existing_t = conn.execute(
                "SELECT id FROM tagihan WHERE no_rekening=? AND bulan=?", (no_rek, bulan)
            ).fetchone()

            if existing_t:
                conn.execute("""
                    UPDATE tagihan SET saldo_pinjaman=?, tunggakan_pokok=?, tunggakan_margin=?,
                    total_tagihan=?, kolektibilitas=? WHERE no_rekening=? AND bulan=?
                """, (saldo, tung_pokok, tung_margin, total, kolek, no_rek, bulan))
                tagihan_update += 1
            else:
                conn.execute("""
                    INSERT INTO tagihan (no_rekening, bulan, saldo_pinjaman, tunggakan_pokok,
                    tunggakan_margin, total_tagihan, kolektibilitas)
                    VALUES (?,?,?,?,?,?,?)
                """, (no_rek, bulan, saldo, tung_pokok, tung_margin, total, kolek))
                tagihan_baru += 1

        # Auto-nonaktifkan nasabah yang tidak ada di Excel bulan ini
        all_nasabah = conn.execute("SELECT no_rekening FROM nasabah WHERE aktif=1").fetchall()
        for row in all_nasabah:
            if row[0] not in no_rek_excel:
                conn.execute("UPDATE nasabah SET aktif=0 WHERE no_rekening=?", (row[0],))
                nasabah_nonaktif += 1

        # Catat log import
        conn.execute("""
            INSERT INTO import_log (bulan, filepath, nasabah_baru, nasabah_update, nasabah_nonaktif,
            tagihan_baru, tagihan_update, diimport_oleh)
            VALUES (?,?,?,?,?,?,?,?)
        """, (bulan, os.path.basename(filepath), nasabah_baru, nasabah_update,
              nasabah_nonaktif, tagihan_baru, tagihan_update, diimport_oleh))

        conn.commit()
        conn.close()

        return {
            "success": True,
            "bulan": bulan,
            "nasabah_baru": nasabah_baru,
            "nasabah_update": nasabah_update,
            "nasabah_nonaktif": nasabah_nonaktif,
            "tagihan_baru": tagihan_baru,
            "tagihan_update": tagihan_update
        }

    except Exception as e:
        return {"success": False, "error": str(e)}


if __name__ == "__main__":
    init_db()
    print("DB siap dipakai!")
